import unittest
from datetime import date
from decimal import Decimal
from io import BytesIO
from types import SimpleNamespace
from unittest.mock import MagicMock, patch

from openpyxl import load_workbook

from app.services.autoservice_warehouse_stock_report import (
    WarehouseStockReportFilters,
    _balance,
    _month_bounds,
    build_warehouse_stock_report,
)
from app.services.autoservice_warehouse_stock_report_xlsx import build_warehouse_stock_workbook_bytes


def _item(
    item_id=1,
    *,
    brand="MANN",
    article="W712",
    name="Фильтр",
    unit="pcs",
    unit_price=Decimal("100.00"),
    quantity=5,
    reserved_qty=1,
    return_reserved_qty=0,
):
    return SimpleNamespace(
        id=item_id,
        brand=brand,
        article=article,
        name=name,
        unit=unit,
        unit_price=unit_price,
        quantity=quantity,
        reserved_qty=reserved_qty,
        return_reserved_qty=return_reserved_qty,
    )


class WarehouseStockReportHelpersTests(unittest.TestCase):
    def test_month_bounds(self):
        start, end, opening = _month_bounds(2026, 8)
        self.assertEqual(start, date(2026, 8, 1))
        self.assertEqual(end, date(2026, 8, 31))
        self.assertEqual(opening, date(2026, 7, 31))

    def test_balance_never_negative(self):
        self.assertEqual(_balance(3, 10), 0)
        self.assertEqual(_balance(10, 3), 7)


class WarehouseStockReportBuildTests(unittest.TestCase):
    def _filters(self, **kwargs):
        defaults = {"year": 2026, "month": 8, "hide_zero": True}
        defaults.update(kwargs)
        return WarehouseStockReportFilters(**defaults)

    def _run_report(self, items, aggregates=None, filters=None):
        aggregates = aggregates or {}
        db = MagicMock()

        item_query = MagicMock()
        item_query.filter.return_value.order_by.return_value.all.return_value = items
        db.query.return_value = item_query

        def aggregate_side_effect(db_arg, org_id, model, *, date_to=None, date_from=None):
            key = (model.__name__, date_to, date_from)
            return aggregates.get(key, {})

        with patch(
            "app.services.autoservice_warehouse_stock_report._aggregate_qty_by_item",
            side_effect=aggregate_side_effect,
        ):
            with patch(
                "app.services.autoservice_warehouse_stock_report._is_current_month",
                return_value=aggregates.get("is_current_month", False),
            ):
                return build_warehouse_stock_report(
                    db,
                    "ORG1",
                    filters or self._filters(),
                )

    def test_closing_qty_from_movements(self):
        from app.models.autoservice_warehouse import (
            AutoserviceWarehouseExpense,
            AutoserviceWarehouseReceipt,
        )

        month_start, month_end, opening_as_of = _month_bounds(2026, 8)
        aggregates = {
            (AutoserviceWarehouseReceipt.__name__, opening_as_of, None): {1: 10},
            (AutoserviceWarehouseExpense.__name__, opening_as_of, None): {1: 2},
            (AutoserviceWarehouseReceipt.__name__, month_end, None): {1: 15},
            (AutoserviceWarehouseExpense.__name__, month_end, None): {1: 4},
            (AutoserviceWarehouseReceipt.__name__, month_end, month_start): {1: 5},
            (AutoserviceWarehouseExpense.__name__, month_end, month_start): {1: 2},
        }
        report = self._run_report([_item()], aggregates=aggregates)
        row = report["items"][0]
        self.assertEqual(row["opening_qty"], 8)
        self.assertEqual(row["closing_qty"], 11)
        self.assertEqual(row["received_qty"], 5)
        self.assertEqual(row["expensed_qty"], 2)
        self.assertEqual(row["stock_amount"], Decimal("1100.00"))

    def test_hide_zero_excludes_zero_closing(self):
        from app.models.autoservice_warehouse import (
            AutoserviceWarehouseExpense,
            AutoserviceWarehouseReceipt,
        )

        month_start, month_end, opening_as_of = _month_bounds(2026, 8)
        aggregates = {
            (AutoserviceWarehouseReceipt.__name__, opening_as_of, None): {1: 0, 2: 3},
            (AutoserviceWarehouseExpense.__name__, opening_as_of, None): {},
            (AutoserviceWarehouseReceipt.__name__, month_end, None): {1: 0, 2: 3},
            (AutoserviceWarehouseExpense.__name__, month_end, None): {1: 0, 2: 0},
            (AutoserviceWarehouseReceipt.__name__, month_end, month_start): {},
            (AutoserviceWarehouseExpense.__name__, month_end, month_start): {},
        }
        report = self._run_report(
            [_item(1), _item(2, brand="BOSCH", article="A1", name="Свеча")],
            aggregates=aggregates,
        )
        self.assertEqual(len(report["items"]), 1)
        self.assertEqual(report["items"][0]["id"], 2)

    def test_search_filters_by_brand(self):
        from app.models.autoservice_warehouse import (
            AutoserviceWarehouseExpense,
            AutoserviceWarehouseReceipt,
        )

        month_start, month_end, opening_as_of = _month_bounds(2026, 8)
        aggregates = {
            (AutoserviceWarehouseReceipt.__name__, opening_as_of, None): {1: 1, 2: 1},
            (AutoserviceWarehouseExpense.__name__, opening_as_of, None): {},
            (AutoserviceWarehouseReceipt.__name__, month_end, None): {1: 1, 2: 1},
            (AutoserviceWarehouseExpense.__name__, month_end, None): {},
            (AutoserviceWarehouseReceipt.__name__, month_end, month_start): {},
            (AutoserviceWarehouseExpense.__name__, month_end, month_start): {},
        }
        report = self._run_report(
            [_item(1), _item(2, brand="BOSCH", article="A1", name="Свеча")],
            aggregates=aggregates,
            filters=self._filters(q="bosch"),
        )
        self.assertEqual(len(report["items"]), 1)
        self.assertEqual(report["items"][0]["brand"], "BOSCH")

    def test_current_month_includes_available_qty(self):
        from app.models.autoservice_warehouse import (
            AutoserviceWarehouseExpense,
            AutoserviceWarehouseReceipt,
        )

        month_start, month_end, opening_as_of = _month_bounds(2026, 8)
        aggregates = {
            "is_current_month": True,
            (AutoserviceWarehouseReceipt.__name__, opening_as_of, None): {1: 5},
            (AutoserviceWarehouseExpense.__name__, opening_as_of, None): {},
            (AutoserviceWarehouseReceipt.__name__, month_end, None): {1: 5},
            (AutoserviceWarehouseExpense.__name__, month_end, None): {},
            (AutoserviceWarehouseReceipt.__name__, month_end, month_start): {},
            (AutoserviceWarehouseExpense.__name__, month_end, month_start): {},
        }
        with patch(
            "app.services.autoservice_warehouse_stock_report.autoservice_item_available_qty",
            return_value=4,
        ):
            report = self._run_report([_item(reserved_qty=1)], aggregates=aggregates)
        row = report["items"][0]
        self.assertTrue(report["is_current_month"])
        self.assertEqual(row["reserved_qty"], 1)
        self.assertEqual(row["available_qty"], 4)

    def test_past_month_has_no_reserved_fields(self):
        from app.models.autoservice_warehouse import (
            AutoserviceWarehouseExpense,
            AutoserviceWarehouseReceipt,
        )

        month_start, month_end, opening_as_of = _month_bounds(2026, 7)
        aggregates = {
            "is_current_month": False,
            (AutoserviceWarehouseReceipt.__name__, opening_as_of, None): {1: 2},
            (AutoserviceWarehouseExpense.__name__, opening_as_of, None): {},
            (AutoserviceWarehouseReceipt.__name__, month_end, None): {1: 2},
            (AutoserviceWarehouseExpense.__name__, month_end, None): {},
            (AutoserviceWarehouseReceipt.__name__, month_end, month_start): {},
            (AutoserviceWarehouseExpense.__name__, month_end, month_start): {},
        }
        report = self._run_report(
            [_item()],
            aggregates=aggregates,
            filters=self._filters(year=2026, month=7),
        )
        row = report["items"][0]
        self.assertFalse(report["is_current_month"])
        self.assertIsNone(row["reserved_qty"])
        self.assertIsNone(row["available_qty"])


class WarehouseStockReportXlsxTests(unittest.TestCase):
    def test_workbook_is_non_empty(self):
        db = MagicMock()
        filters = WarehouseStockReportFilters(year=2026, month=8)
        with patch(
            "app.services.autoservice_warehouse_stock_report_xlsx.build_warehouse_stock_report",
            return_value={
                "year": 2026,
                "month": 8,
                "as_of": date(2026, 8, 31),
                "is_current_month": True,
                "summary": {
                    "positions": 1,
                    "closing_value": Decimal("100.00"),
                    "opening_value": Decimal("50.00"),
                    "received_qty": 2,
                    "expensed_qty": 1,
                },
                "items": [{
                    "id": 1,
                    "brand": "MANN",
                    "article": "W712",
                    "name": "Фильтр",
                    "unit": "pcs",
                    "unit_price": Decimal("100.00"),
                    "opening_qty": 1,
                    "received_qty": 2,
                    "expensed_qty": 1,
                    "closing_qty": 2,
                    "reserved_qty": 0,
                    "return_reserved_qty": 0,
                    "available_qty": 2,
                    "stock_amount": Decimal("200.00"),
                }],
            },
        ):
            content = build_warehouse_stock_workbook_bytes(db, "ORG1", filters)
        self.assertTrue(content)
        wb = load_workbook(BytesIO(content))
        self.assertEqual(wb.sheetnames, ["Сводка", "Остатки"])


if __name__ == "__main__":
    unittest.main()
