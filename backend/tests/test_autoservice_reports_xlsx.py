import unittest
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from unittest.mock import MagicMock, patch

from openpyxl import load_workbook

from app.schemas.autoservice_finance import (
    AutoserviceFinanceReceiptRow,
    AutoserviceFinanceReceiptsResponse,
    AutoservicePaymentMethodTotals,
)
from app.services.autoservice_finance_receipts_xlsx import build_finance_receipts_workbook_bytes
from app.services.autoservice_payroll_report_xlsx import build_payroll_report_workbook_bytes


class FinanceReceiptsXlsxTests(unittest.TestCase):
    def test_workbook_is_non_empty(self):
        db = MagicMock()
        report = AutoserviceFinanceReceiptsResponse(
            totals=AutoservicePaymentMethodTotals(
                card=Decimal("100.00"),
                cash=Decimal("50.00"),
                bank=Decimal("0.00"),
            ),
            total_amount=Decimal("150.00"),
            count=2,
            items=[
                AutoserviceFinanceReceiptRow(
                    id=1,
                    sequential_number=1,
                    repair_order_id=10,
                    repair_order_number="ЗН-001",
                    client_name="Иван",
                    amount=Decimal("100.00"),
                    method="card",
                    created_at=datetime(2026, 8, 15, 12, 0, 0),
                ),
                AutoserviceFinanceReceiptRow(
                    id=2,
                    sequential_number=2,
                    repair_order_id=11,
                    repair_order_number="ЗН-002",
                    client_name="Пётр",
                    amount=Decimal("50.00"),
                    method="cash",
                    created_at=datetime(2026, 8, 16, 10, 30, 0),
                ),
            ],
        )
        with patch(
            "app.services.autoservice_finance_receipts_xlsx.list_finance_receipts",
            return_value=report,
        ):
            content = build_finance_receipts_workbook_bytes(
                db,
                "ORG1",
                date(2026, 8, 1),
                date(2026, 8, 31),
            )
        self.assertTrue(content)
        wb = load_workbook(BytesIO(content))
        self.assertEqual(wb.sheetnames, ["Сводка", "Платежи"])
        self.assertEqual(wb["Сводка"]["A1"].value, "Платежи автосервиса")
        self.assertEqual(wb["Платежи"]["A1"].value, "№")
        self.assertEqual(wb["Платежи"]["C1"].value, "Клиент")
        self.assertEqual(wb["Платежи"]["C2"].value, "Иван")
        self.assertEqual(wb["Платежи"]["C3"].value, "Пётр")


class PayrollReportXlsxTests(unittest.TestCase):
    def test_workbook_is_non_empty(self):
        db = MagicMock()
        report = {
            "year": 2026,
            "month": 8,
            "total": Decimal("300.00"),
            "employees": [
                {
                    "employee_id": 1,
                    "name": "Мастер",
                    "completed_orders": 1,
                    "total": Decimal("300.00"),
                    "orders": [
                        {
                            "order_id": 10,
                            "order_number": "ЗН-001",
                            "vehicle": {
                                "id": 1,
                                "make": "Toyota",
                                "model": "Camry",
                                "year": 2020,
                                "plate": "A123BC",
                            },
                            "amount": Decimal("300.00"),
                            "works": [
                                {
                                    "work_id": 501,
                                    "title": "Замена масла",
                                    "qty": 1,
                                    "unit_price": Decimal("1000.00"),
                                    "line_total": Decimal("1000.00"),
                                    "percent": Decimal("30.00"),
                                    "accrual_type": "work_percent",
                                    "accrual_type_label": "Процент от работы",
                                    "amount": Decimal("300.00"),
                                }
                            ],
                        }
                    ],
                }
            ],
        }
        with patch(
            "app.services.autoservice_payroll_report_xlsx.compute_org_monthly_payroll",
            return_value=report,
        ):
            content = build_payroll_report_workbook_bytes(db, "ORG1", 2026, 8)
        self.assertTrue(content)
        wb = load_workbook(BytesIO(content))
        self.assertEqual(wb.sheetnames, ["Сводка", "Сотрудники", "По заказ-нарядам", "Работы"])
        self.assertEqual(wb["Сводка"]["A1"].value, "Зарплаты автосервиса")
        self.assertEqual(wb["Работы"]["D1"].value, "Работа")
        self.assertEqual(wb["Работы"]["D2"].value, "Замена масла")


if __name__ == "__main__":
    unittest.main()
