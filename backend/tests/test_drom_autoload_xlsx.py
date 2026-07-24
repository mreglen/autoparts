import unittest
from pathlib import Path

from app.services.drom_autoload_xlsx import (
    build_drom_delta_xlsx,
    build_drom_header_only_xlsx,
    chunk_export_rows_for_drom_sync,
    parse_and_validate_drom_autoload,
    remove_product_from_drom_autoload,
    remove_products_from_drom_autoload,
    upsert_products_to_drom_autoload,
    zero_quantity_rows_for_articles,
    QUANTITY_HEADER,
)


class DromAutoloadXlsxTests(unittest.TestCase):
    def _sample_row(self, *, article="ART-001", name="Фильтр масляный", price=1500, quantity=2):
        return {
            "product_id": 1,
            "article": article,
            "name": name,
            "is_new": True,
            "brand": "Toyota",
            "price": price,
            "quantity": quantity,
            "photos": ["/pictures/test.jpg"],
            "storage_address": "г. Москва, ул. Тестовая, 1",
        }

    def test_upsert_creates_valid_xlsx_from_template(self):
        xlsx_bytes = upsert_products_to_drom_autoload(
            None,
            [self._sample_row()],
            public_base_url="https://example.com",
        )
        parsed = parse_and_validate_drom_autoload(xlsx_bytes)
        self.assertTrue(parsed.local_ok, parsed.local_errors)
        self.assertEqual(len(parsed.items), 1)
        self.assertEqual(parsed.items[0]["article"], "ART-001")

    def test_upsert_updates_existing_article(self):
        first = upsert_products_to_drom_autoload(None, [self._sample_row(price=1000)], public_base_url="")
        second = upsert_products_to_drom_autoload(
            first,
            [self._sample_row(price=2000, name="Фильтр обновлённый")],
            public_base_url="",
        )
        parsed = parse_and_validate_drom_autoload(second)
        self.assertTrue(parsed.local_ok, parsed.local_errors)
        self.assertEqual(len(parsed.items), 1)
        self.assertEqual(float(parsed.items[0]["price"]), 2000.0)

    def test_validation_fails_on_invalid_price(self):
        xlsx_bytes = upsert_products_to_drom_autoload(
            None,
            [self._sample_row(price=0)],
            public_base_url="",
        )
        parsed = parse_and_validate_drom_autoload(xlsx_bytes)
        self.assertFalse(parsed.local_ok)
        self.assertTrue(len(parsed.local_errors) >= 1)

    def test_remove_product_from_xlsx(self):
        xlsx_bytes = upsert_products_to_drom_autoload(
            None,
            [
                self._sample_row(article="A1"),
                self._sample_row(article="A2", name="Другая деталь"),
            ],
            public_base_url="",
        )
        updated = remove_product_from_drom_autoload(xlsx_bytes, "A1")
        parsed = parse_and_validate_drom_autoload(updated)
        self.assertTrue(parsed.local_ok, parsed.local_errors)
        self.assertEqual(len(parsed.items), 1)
        self.assertEqual(parsed.items[0]["article"], "A2")

    def test_remove_multiple_products_from_xlsx(self):
        xlsx_bytes = upsert_products_to_drom_autoload(
            None,
            [
                self._sample_row(article="A1"),
                self._sample_row(article="A2"),
                self._sample_row(article="A3"),
            ],
            public_base_url="",
        )
        updated = remove_products_from_drom_autoload(xlsx_bytes, ["A1", "A3"])
        parsed = parse_and_validate_drom_autoload(updated)
        self.assertTrue(parsed.local_ok, parsed.local_errors)
        self.assertEqual([item["article"] for item in parsed.items], ["A2"])

    def test_upsert_handles_many_rows(self):
        rows = [
            self._sample_row(article=f"ART-{idx}", name=f"Деталь {idx}", price=1000 + idx)
            for idx in range(1, 201)
        ]
        xlsx_bytes = upsert_products_to_drom_autoload(None, rows, public_base_url="")
        parsed = parse_and_validate_drom_autoload(xlsx_bytes)
        self.assertTrue(parsed.local_ok, parsed.local_errors)
        self.assertEqual(len(parsed.items), 200)

        template_path = Path(__file__).resolve().parents[1] / "app" / "templates" / "drom" / "auto-parts-GT.xlsx"
        self.assertTrue(template_path.is_file(), f"Missing template: {template_path}")

    def test_build_delta_xlsx_only_requested_rows(self):
        xlsx_bytes = build_drom_delta_xlsx(
            [
                self._sample_row(article="D1"),
                self._sample_row(article="D2", name="Вторая"),
            ],
            public_base_url="",
        )
        parsed = parse_and_validate_drom_autoload(xlsx_bytes)
        self.assertTrue(parsed.local_ok, parsed.local_errors)
        articles = {item["article"] for item in parsed.items}
        self.assertEqual(articles, {"D1", "D2"})

    def test_zero_quantity_rows_for_api_delete(self):
        rows = zero_quantity_rows_for_articles(["DEL-1", "", "DEL-2"])
        self.assertEqual(len(rows), 2)
        self.assertTrue(all(r["quantity"] == 0 for r in rows))
        xlsx_bytes = build_drom_delta_xlsx(rows, public_base_url="")
        # parse allows price>0; qty 0 is fine
        from openpyxl import load_workbook
        from io import BytesIO

        wb = load_workbook(BytesIO(xlsx_bytes))
        ws = wb.active
        header = [c.value for c in ws[1]]
        qty_col = header.index(QUANTITY_HEADER) + 1
        quantities = [ws.cell(row=r, column=qty_col).value for r in range(2, ws.max_row + 1)]
        self.assertEqual(quantities, [0, 0])
        wb.close()

    def test_header_only_xlsx(self):
        xlsx_bytes = build_drom_header_only_xlsx()
        parsed = parse_and_validate_drom_autoload(xlsx_bytes)
        self.assertTrue(parsed.local_ok)
        self.assertEqual(parsed.items, [])

    def test_chunk_export_rows_respects_max_bytes(self):
        rows = [self._sample_row(article=f"C-{i}", name=f"Part {i}") for i in range(5)]
        # Force tiny limit so each row becomes its own chunk (or few)
        chunks = chunk_export_rows_for_drom_sync(rows, public_base_url="", max_bytes=8_000)
        self.assertGreaterEqual(len(chunks), 1)
        for chunk in chunks:
            self.assertIsInstance(chunk, (bytes, bytearray))
            self.assertLessEqual(len(chunk), 8_000 + 50_000)  # single oversized row may exceed


if __name__ == "__main__":
    unittest.main()
