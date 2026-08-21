import unittest
from datetime import datetime
from decimal import Decimal
from io import BytesIO
from types import SimpleNamespace
from unittest.mock import MagicMock, patch

from openpyxl import load_workbook

from app.services.autoservice_order_economics import (
    OrderEconomicsFilters,
    _build_order_row,
    _money,
    _order_grand_total,
    _order_parts_cost,
    _payment_status,
    _preview_order_payroll,
    build_order_economics_report,
    order_economics_payment_label,
    order_economics_status_label,
)
from app.services.autoservice_order_economics_xlsx import build_order_economics_workbook_bytes


def _client():
    return SimpleNamespace(id=1, name="Иванов И.", phone="+79990001122")


def _vehicle():
    return SimpleNamespace(id=10, make="Toyota", model="Camry", year=2020, plate="A123BC")


def _executor(employee_id=5, percent=Decimal("50.00")):
    employee = SimpleNamespace(id=employee_id, name="Петров")
    return SimpleNamespace(employee_id=employee_id, percent=percent, employee=employee)


def _work(work_id=1, qty=1, unit_price=Decimal("1000.00"), executors=None):
    return SimpleNamespace(
        id=work_id,
        position=work_id,
        qty=qty,
        unit_price=unit_price,
        executors=executors or [_executor()],
    )


def _shop_part(part_id=1, qty=Decimal("1.000"), unit_price=Decimal("500.00"), markup_percent=Decimal("20.00")):
    return SimpleNamespace(
        id=part_id,
        position=part_id,
        qty=qty,
        unit="pcs",
        unit_price=unit_price,
        markup_percent=markup_percent,
        client_unit_price_override=None,
        source="manual",
    )


def _order(
    order_id=100,
    status="completed",
    works=None,
    shop_parts=None,
    scheduled_at=None,
):
    return SimpleNamespace(
        id=order_id,
        order_number="RO-100",
        status=status,
        organization_id="ORG1",
        client_id=1,
        vehicle_id=10,
        scheduled_at=scheduled_at or datetime(2026, 8, 15, 10, 0),
        client=_client(),
        vehicle=_vehicle(),
        works=works if works is not None else [_work()],
        shop_parts=shop_parts if shop_parts is not None else [_shop_part()],
    )


class AutoserviceOrderEconomicsTests(unittest.TestCase):
    def _filters(self, **kwargs):
        defaults = {
            "date_from": datetime(2026, 8, 1).date(),
            "date_to": datetime(2026, 8, 31).date(),
        }
        defaults.update(kwargs)
        return OrderEconomicsFilters(**defaults)

    def _run_report(self, orders, *, accruals=None, employees=None, paid_map=None, filters=None):
        db = MagicMock()
        query = MagicMock()
        query.filter.return_value = query
        query.order_by.return_value = query
        query.all.return_value = orders

        accruals = accruals or []
        employees = employees or []
        paid_map = paid_map or {}

        payroll_rows = {}
        for accrual in accruals:
            payroll_rows[accrual.order_id] = payroll_rows.get(accrual.order_id, Decimal("0.00")) + accrual.amount

        def query_side_effect(*args, **kwargs):
            q = MagicMock()
            q.filter.return_value = q
            q.group_by.return_value = q
            model = args[0] if args else None
            model_name = getattr(model, "__name__", "")
            if model_name == "AutoserviceServiceEmployee":
                q.all.return_value = employees
            else:
                q.all.return_value = list(payroll_rows.items()) if payroll_rows else []
            return q

        db.query.side_effect = query_side_effect
        filters = filters or self._filters()

        payroll_map = {order_id: _money(amount) for order_id, amount in payroll_rows.items()}

        with patch("app.services.autoservice_order_economics._order_economics_query", return_value=query), patch(
            "app.services.autoservice_order_economics.batch_paid_amounts",
            return_value={order.id: paid_map.get(order.id, Decimal("0.00")) for order in orders},
        ), patch(
            "app.services.autoservice_order_economics._batch_payroll_totals",
            return_value=payroll_map,
        ):
            return build_order_economics_report(db, "ORG1", filters)

    def test_order_totals_and_preview_payroll(self):
        order = _order(
            works=[_work(unit_price=Decimal("1000.00"), executors=[_executor(percent=Decimal("10.00"))])],
            shop_parts=[_shop_part(unit_price=Decimal("400.00"), markup_percent=Decimal("25.00"))],
        )
        self.assertEqual(_order_grand_total(order), Decimal("1500.00"))
        self.assertEqual(_order_parts_cost(order), Decimal("400.00"))

        daily_employee = SimpleNamespace(
            id=5,
            salary_type="daily_rate",
            salary_amount=Decimal("1500.00"),
        )
        preview = _preview_order_payroll(order, {5: daily_employee})
        self.assertEqual(preview, Decimal("1600.00"))

    def test_completed_order_uses_accruals_and_calculates_profit(self):
        order = _order(
            works=[_work(unit_price=Decimal("1000.00"), executors=[_executor(percent=Decimal("10.00"))])],
            shop_parts=[_shop_part(unit_price=Decimal("400.00"), markup_percent=Decimal("25.00"))],
        )
        result = self._run_report(
            [order],
            accruals=[SimpleNamespace(order_id=100, amount=Decimal("100.00"))],
            paid_map={100: Decimal("1500.00")},
        )

        self.assertEqual(result["summary"]["count"], 1)
        self.assertEqual(result["summary"]["revenue"], Decimal("1500.00"))
        self.assertEqual(result["summary"]["parts_cost"], Decimal("400.00"))
        self.assertEqual(result["summary"]["payroll_total"], Decimal("100.00"))
        self.assertEqual(result["summary"]["net_profit"], Decimal("1000.00"))
        row = result["items"][0]
        self.assertFalse(row["is_preliminary"])
        self.assertEqual(row["payment_status"], "paid")

    def test_open_order_uses_preview_payroll(self):
        daily_employee = SimpleNamespace(
            id=5,
            name="Петров",
            salary_type="daily_rate",
            salary_amount=Decimal("1500.00"),
            organization_id="ORG1",
        )
        order = _order(status="in_progress")
        result = self._run_report([order], employees=[daily_employee], paid_map={100: Decimal("0.00")})
        row = result["items"][0]

        self.assertTrue(row["is_preliminary"])
        self.assertEqual(row["payroll_total"], Decimal("2000.00"))
        self.assertEqual(row["payment_status"], "unpaid")

    def test_payment_filter_partial(self):
        paid_order = _order(order_id=101, status="done")
        paid_order.order_number = "RO-101"
        partial_order = _order(order_id=102, status="done")
        partial_order.order_number = "RO-102"
        result = self._run_report(
            [paid_order, partial_order],
            paid_map={101: Decimal("1600.00"), 102: Decimal("500.00")},
            filters=self._filters(payment="partial"),
        )

        self.assertEqual(result["summary"]["count"], 1)
        self.assertEqual(result["items"][0]["order_id"], 102)
        self.assertEqual(result["items"][0]["payment_status"], "partial")

    def test_status_filter(self):
        pending = _order(order_id=201, status="pending")
        cancelled = _order(order_id=202, status="cancelled")
        result = self._run_report(
            [pending, cancelled],
            filters=self._filters(status="pending"),
        )
        self.assertEqual(len(result["items"]), 1)
        self.assertEqual(result["items"][0]["order_id"], 201)

    def test_payment_status_helper(self):
        self.assertEqual(_payment_status(Decimal("1500.00"), Decimal("0.00")), "paid")
        self.assertEqual(_payment_status(Decimal("500.00"), Decimal("1000.00")), "partial")
        self.assertEqual(_payment_status(Decimal("0.00"), Decimal("1500.00")), "unpaid")

    def test_build_order_row(self):
        order = _order(status="done")
        row = _build_order_row(
            order,
            paid_amount=Decimal("0.00"),
            payroll_total=Decimal("500.00"),
            is_preliminary=True,
        )
        self.assertTrue(row["is_preliminary"])
        self.assertEqual(row["client_name"], "Иванов И.")
        self.assertEqual(row["vehicle"]["plate"], "A123BC")

    def test_labels(self):
        self.assertEqual(order_economics_status_label("completed"), "Закрыт")
        self.assertEqual(order_economics_payment_label("unpaid"), "Долг")

    def test_xlsx_workbook_has_summary_and_orders_sheets(self):
        order = _order()
        db = MagicMock()
        filters = self._filters()
        report = {
            "date_from": filters.date_from,
            "date_to": filters.date_to,
            "summary": {
                "count": 1,
                "revenue": Decimal("1500.00"),
                "parts_cost": Decimal("400.00"),
                "payroll_total": Decimal("100.00"),
                "net_profit": Decimal("1000.00"),
                "paid_amount": Decimal("0.00"),
                "debt_amount": Decimal("1500.00"),
                "unpaid_count": 1,
            },
            "items": [
                _build_order_row(
                    order,
                    paid_amount=Decimal("0.00"),
                    payroll_total=Decimal("100.00"),
                    is_preliminary=False,
                )
            ],
        }
        with patch(
            "app.services.autoservice_order_economics_xlsx.build_order_economics_report",
            return_value=report,
        ):
            content = build_order_economics_workbook_bytes(db, "ORG1", filters)

        wb = load_workbook(BytesIO(content))
        self.assertEqual(wb.sheetnames, ["Сводка", "Заказ-наряды"])
        self.assertEqual(wb["Сводка"]["A1"].value, "Сводная таблица")
        self.assertEqual(wb["Заказ-наряды"]["A1"].value, "Заказ-наряд")
