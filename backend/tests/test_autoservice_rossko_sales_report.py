import unittest
from datetime import datetime, timezone
from decimal import Decimal
from io import BytesIO
from types import SimpleNamespace
from unittest.mock import MagicMock, patch

from fastapi import HTTPException
from openpyxl import load_workbook

from app.services.autoservice_rossko_sales_report import RosskoSalesReportFilters, build_rossko_sales_report
from app.services.autoservice_rossko_sales_report_xlsx import build_rossko_sales_workbook_bytes
from app.utils.autoservice_access import organization_has_admin_director, require_rossko_sales_report_access


def _order(*, order_id=1, paid=True, yk_id="pay-1", session_id=None):
    item = SimpleNamespace(
        id=11,
        order_id=order_id,
        brand="BOSCH",
        partnumber="123",
        name="Filter",
        quantity=2,
        price=500.0,
        supplier_unit_price=350.0,
    )
    return SimpleNamespace(
        id=order_id,
        organization_id="ORG1",
        buyer_name="Иван",
        buyer_phone="+79990000000",
        total_amount=1000.0,
        is_paid=paid,
        rossko_order_id="RK-1",
        checkout_session_id=session_id,
        yookassa_payment_id=yk_id if paid else None,
        created_at=datetime(2026, 8, 10, 12, 0, tzinfo=timezone.utc),
        items=[item],
    )


class RosskoSalesReportTests(unittest.TestCase):
    def test_build_report_includes_paid_order(self):
        order = _order()
        payment = SimpleNamespace(
            id="p1",
            session_id="s1",
            yookassa_payment_id="pay-1",
            payment_method_type="sbp",
            status="succeeded",
            amount_value=1000.0,
            income_amount=970.0,
            acquiring_fee_amount=30.0,
            paid_at=datetime(2026, 8, 10, 12, 5, tzinfo=timezone.utc),
            refund_status=None,
            refund_amount=None,
            refunded_at=None,
        )
        db = MagicMock()
        with patch(
            "app.services.autoservice_rossko_sales_report._fetch_report_orders",
            return_value=[order],
        ), patch(
            "app.services.autoservice_rossko_sales_report._load_payment_map",
            return_value={"pay-1": payment},
        ):
            report = build_rossko_sales_report(
                db,
                "ORG1",
                RosskoSalesReportFilters(
                    date_from=datetime(2026, 8, 1).date(),
                    date_to=datetime(2026, 8, 31).date(),
                ),
            )
        self.assertEqual(report["summary"]["count"], 1)
        self.assertEqual(report["summary"]["sale_total"], Decimal("1000.00"))
        self.assertEqual(report["summary"]["site_income"], Decimal("18.90"))
        row = report["items"][0]
        self.assertEqual(row["payment_method"], "sbp")
        self.assertEqual(len(row["items"]), 1)

    def test_build_report_includes_unpaid_order_by_created_at(self):
        order = _order(paid=False, yk_id=None, session_id=None)
        db = MagicMock()
        with patch(
            "app.services.autoservice_rossko_sales_report._fetch_report_orders",
            return_value=[order],
        ), patch(
            "app.services.autoservice_rossko_sales_report._load_payment_map",
            return_value={},
        ):
            report = build_rossko_sales_report(
                db,
                "ORG1",
                RosskoSalesReportFilters(
                    date_from=datetime(2026, 8, 1).date(),
                    date_to=datetime(2026, 8, 31).date(),
                ),
            )
        self.assertEqual(report["summary"]["count"], 1)
        self.assertEqual(report["items"][0]["payment_method"], "unpaid")
        self.assertEqual(report["items"][0]["acquiring_fee"], Decimal("0.00"))

    def test_access_requires_admin_director_org(self):
        db = MagicMock()
        user = SimpleNamespace(id=1, organization_id="ORG1", is_employee=True)
        with patch(
            "app.utils.autoservice_access.require_autoservice_permission",
            return_value="ORG1",
        ), patch(
            "app.utils.autoservice_access.organization_has_admin_director",
            return_value=False,
        ):
            with self.assertRaises(HTTPException) as ctx:
                require_rossko_sales_report_access(db, user)
            self.assertEqual(ctx.exception.status_code, 403)

    def test_organization_has_admin_director(self):
        db = MagicMock()
        db.query.return_value.filter.return_value.first.return_value = 1
        self.assertTrue(organization_has_admin_director(db, "ORG1"))


class RosskoSalesXlsxTests(unittest.TestCase):
    def test_workbook_has_expected_sheets(self):
        db = MagicMock()
        report = {
            "date_from": datetime(2026, 8, 1).date(),
            "date_to": datetime(2026, 8, 31).date(),
            "summary": {
                "count": 1,
                "sale_total": Decimal("1000.00"),
                "supplier_total": Decimal("700.00"),
                "acquiring_fee": Decimal("30.00"),
                "refund_total": Decimal("0.00"),
                "margin": Decimal("270.00"),
                "site_income": Decimal("18.90"),
                "organization_income": Decimal("251.10"),
                "pending_count": 0,
            },
            "items": [
                {
                    "order_id": 1,
                    "operation_at": "2026-08-10T12:05:00+00:00",
                    "rossko_order_id": "RK-1",
                    "buyer_name": "Иван",
                    "payment_method_label": "СБП",
                    "sale_total": Decimal("1000.00"),
                    "supplier_total": Decimal("700.00"),
                    "acquiring_fee": Decimal("30.00"),
                    "refund_amount": Decimal("0.00"),
                    "margin": Decimal("270.00"),
                    "site_income": Decimal("18.90"),
                    "organization_income": Decimal("251.10"),
                    "pending_acquiring": False,
                    "items": [
                        {
                            "brand": "BOSCH",
                            "partnumber": "123",
                            "name": "Filter",
                            "quantity": 2,
                            "sale_total": Decimal("1000.00"),
                            "supplier_total": Decimal("700.00"),
                            "acquiring_fee": Decimal("30.00"),
                            "refund_amount": Decimal("0.00"),
                            "margin": Decimal("270.00"),
                            "site_income": Decimal("18.90"),
                            "organization_income": Decimal("251.10"),
                        }
                    ],
                }
            ],
        }
        with patch(
            "app.services.autoservice_rossko_sales_report_xlsx.build_rossko_sales_report",
            return_value=report,
        ):
            content = build_rossko_sales_workbook_bytes(
                db,
                "ORG1",
                RosskoSalesReportFilters(
                    date_from=datetime(2026, 8, 1).date(),
                    date_to=datetime(2026, 8, 31).date(),
                ),
            )
        wb = load_workbook(BytesIO(content))
        self.assertEqual(wb.sheetnames, ["Сводка", "Операции", "Позиции"])
        self.assertEqual(wb["Сводка"]["A1"].value, "Продажи Росско")


if __name__ == "__main__":
    unittest.main()
