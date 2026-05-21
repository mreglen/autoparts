import unittest
from datetime import date
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker

import app.models  # noqa: F401
import app.models.organization_avito_integration  # noqa: F401
import app.models.organization_drom_integration  # noqa: F401
from app.models.product import Product
from app.models.stock_out import StockOut
from app.services.finance_reports import FinanceFilters
from app.services.finance_xlsx_export import build_finance_workbook_bytes


class FinanceXlsxExportTests(unittest.TestCase):
    def setUp(self):
        self.engine = create_engine("sqlite:///:memory:")
        with self.engine.begin() as conn:
            conn.execute(
                text(
                    """
                    CREATE TABLE products (
                        id INTEGER PRIMARY KEY,
                        article VARCHAR(30),
                        name VARCHAR(255),
                        brand VARCHAR(100),
                        internal_code VARCHAR(100) NOT NULL,
                        description TEXT,
                        is_new BOOLEAN,
                        price NUMERIC(12, 2),
                        quantity INTEGER,
                        organization_id VARCHAR,
                        storage_location_id INTEGER,
                        created_by INTEGER NOT NULL,
                        part_type_id INTEGER NOT NULL
                    )
                    """
                )
            )
            conn.execute(
                text(
                    """
                    CREATE TABLE stock_out (
                        id INTEGER PRIMARY KEY,
                        quantity INTEGER,
                        sale_price NUMERIC(12, 2),
                        movement_date DATE,
                        organization_id VARCHAR,
                        storage_location_id INTEGER,
                        product_id INTEGER,
                        acquired_product_id INTEGER,
                        user_id INTEGER,
                        reason TEXT,
                        sale_channel VARCHAR(50),
                        avito_order_id VARCHAR(64),
                        source_kind VARCHAR(32),
                        garage_used_order_item_id INTEGER
                    )
                    """
                )
            )
            conn.execute(
                text(
                    """
                    CREATE TABLE stock_in (
                        id INTEGER PRIMARY KEY,
                        quantity INTEGER,
                        sale_price NUMERIC(12, 2),
                        created_at DATE,
                        organization_id VARCHAR,
                        storage_location_id INTEGER,
                        product_id INTEGER,
                        acquired_product_id INTEGER,
                        created_by INTEGER NOT NULL
                    )
                    """
                )
            )
        self.Session = sessionmaker(bind=self.engine, expire_on_commit=False)
        self.db = self.Session()
        self.db.add(
            Product(
                id=1,
                internal_code="P-1",
                article="A-1",
                name="Test",
                brand="B",
                price=100,
                quantity=5,
                organization_id="org1",
                storage_location_id=1,
                created_by=1,
                part_type_id=1,
            )
        )
        self.db.add(
            StockOut(
                organization_id="org1",
                product_id=1,
                quantity=1,
                sale_price=100,
                movement_date=date(2026, 3, 10),
                storage_location_id=1,
                source_kind="warehouse_manual",
            )
        )
        self.db.commit()

    def tearDown(self):
        self.db.close()
        self.engine.dispose()

    def test_workbook_has_five_sheets(self):
        filters = FinanceFilters(
            date_from=date(2026, 3, 1),
            date_to=date(2026, 3, 31),
            as_of_date=date(2026, 3, 31),
        )
        data = build_finance_workbook_bytes(self.db, "org1", filters)
        self.assertGreater(len(data), 100)
        wb = load_workbook(BytesIO(data))
        self.assertEqual(
            set(wb.sheetnames),
            {"Сводка", "Продажи", "Списания", "Поступления", "Остатки"},
        )
