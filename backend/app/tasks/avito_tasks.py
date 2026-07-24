from __future__ import annotations

import asyncio
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.orm import selectinload
from app.celery_app import celery_app
from app.core.config import settings
from app.db.database import SessionLocal
from app.models.organization import Organization as OrganizationModel
from app.models.avito_autoload_job import AvitoAutoloadJob
from app.models.organization_avito_autoload_cache import OrganizationAvitoAutoloadCache
from app.models.organization_avito_integration import OrganizationAvitoIntegration
from app.models.product import Product as ProductModel, ProductPhoto
from app.models.product_avito_listing_link import ProductAvitoListingLink
from app.models.storage_location import StorageLocation as StorageLocationModel
from app.services import avito_api as avito_api_svc
from app.services.avito_autoload_xlsx import parse_and_validate_avito_autoload, upsert_products_to_avito_autoload
from app.services.marketplace_site_footer import append_marketplace_site_info
from app.services.avito_media import product_photo_urls_for_avito_export
from app.utils.avito_crypto import decrypt_secret
from app.services.avito_pro_status_service import is_avito_pro_active


def _json_loads(raw: str | None, default: Any) -> Any:
    if not raw:
        return default
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        return default


def _set_job_state(
    db,
    job_id: int,
    *,
    status: str | None = None,
    stage: str | None = None,
    processed: int | None = None,
    total: int | None = None,
    result_file_ref: str | None = None,
    result: dict[str, Any] | None = None,
    error_summary: str | None = None,
) -> None:
    job = db.query(AvitoAutoloadJob).filter(AvitoAutoloadJob.id == job_id).first()
    if not job:
        return
    if status is not None:
        job.status = status
    if stage is not None:
        job.stage = stage
    if processed is not None:
        job.processed_count = processed
    if total is not None:
        job.total_count = total
    if result_file_ref is not None:
        job.result_file_ref = result_file_ref
    if result is not None:
        job.result_json = json.dumps(result, ensure_ascii=False)
    if error_summary is not None:
        job.error_summary = error_summary
    db.commit()


def _save_autoload_cache(
    db,
    org_id: str,
    *,
    saved_path: str,
    items: list,
    local_validation_ok: bool,
    local_errors: list,
    sheets_parsed: list,
    avito_upload: Any,
    avito_upload_status: int | None,
    avito_report: Any,
    avito_token_error: str | None,
) -> None:
    row = db.query(OrganizationAvitoAutoloadCache).filter(
        OrganizationAvitoAutoloadCache.organization_id == org_id
    ).first()
    if not row:
        row = OrganizationAvitoAutoloadCache(organization_id=org_id)
        db.add(row)
    row.saved_path = saved_path
    row.items_json = json.dumps(items, ensure_ascii=False)
    row.local_validation_ok = local_validation_ok
    row.local_errors_json = json.dumps(local_errors, ensure_ascii=False)
    row.sheets_parsed_json = json.dumps(sheets_parsed, ensure_ascii=False)
    row.avito_upload_json = json.dumps(avito_upload, ensure_ascii=False) if avito_upload is not None else None
    row.avito_upload_status = avito_upload_status
    row.avito_report_json = json.dumps(avito_report, ensure_ascii=False) if avito_report is not None else None
    row.avito_token_error = avito_token_error
    db.commit()


def _resolve_saved_autoload_file(org_id: str) -> tuple[Path, str]:
    base_dir = Path(__file__).resolve().parents[2] / "uploads" / "avito" / org_id
    base_dir.mkdir(parents=True, exist_ok=True)
    for p in base_dir.iterdir():
        if p.is_file() and p.suffix.lower() == ".xlsx" and p.name != "autoload.xlsx":
            try:
                p.unlink()
            except OSError:
                pass
    xlsx_path = base_dir / "autoload.xlsx"
    rel_path = f"/uploads/avito/{org_id}/autoload.xlsx"
    return xlsx_path, rel_path


def _resolve_saved_path_from_cache(db, org_id: str) -> tuple[Path, str] | tuple[None, None]:
    cache = db.query(OrganizationAvitoAutoloadCache).filter(
        OrganizationAvitoAutoloadCache.organization_id == org_id
    ).first()
    if not cache or not cache.saved_path:
        return None, None
    rel_path = cache.saved_path
    xlsx_path = Path(__file__).resolve().parents[2] / rel_path.lstrip("/")
    if not xlsx_path.is_file():
        return None, None
    return xlsx_path, rel_path


def _map_avito_category(product: ProductModel) -> str:
    name = f"{product.name or ''} {product.article or ''}".lower()
    if any(token in name for token in ("двигател", "мотор")):
        return "Запчасти/Двигатель"
    if any(token in name for token in ("кпп", "коробк", "трансмис")):
        return "Запчасти/Трансмиссия"
    if any(token in name for token in ("бампер", "крыло", "двер", "капот")):
        return "Запчасти/Кузов"
    return "Запчасти/Другое"


def _choose_template_sheet(template_bytes: bytes | None, category_name: str) -> str | None:
    if not template_bytes:
        return None
    try:
        from io import BytesIO
        wb = load_workbook(BytesIO(template_bytes), read_only=True, data_only=True)
        candidates = [s for s in wb.sheetnames if s and s.lower() not in ("инструкция", "instruction")]
        wb.close()
        if not candidates:
            return None
        category_low = (category_name or "").lower()
        for sheet_name in candidates:
            if sheet_name.lower() in category_low or category_low in sheet_name.lower():
                return sheet_name
        return candidates[0]
    except Exception:
        return None


@celery_app.task(bind=True, max_retries=3)
def run_avito_export_job(self, job_id: int):
    db = SessionLocal()
    try:
        job = db.query(AvitoAutoloadJob).filter(AvitoAutoloadJob.id == job_id).first()
        if not job:
            return {"status": "failed", "error": "job not found"}

        payload = _json_loads(job.payload_json, {})
        org_id = job.organization_id
        if not is_avito_pro_active(db, org_id):
            _set_job_state(
                db,
                job_id,
                status="failed",
                stage="failed",
                error_summary="Подписка Avito Pro истекла или нет доступа к API Avito",
            )
            return {"status": "failed", "error": "avito pro inactive"}
        product_ids = [int(x) for x in (payload.get("product_ids") or []) if str(x).isdigit()]
        # Автоотправка в Avito отключена по бизнес-правилу:
        # экспорт только формирует/обновляет XLSX файл.
        publish_after_export = False

        _set_job_state(db, job_id, status="processing", stage="load_products", total=len(product_ids), processed=0)
        if not product_ids:
            _set_job_state(db, job_id, status="failed", stage="failed", error_summary="Нет product_ids")
            return {"status": "failed", "error": "empty product ids"}

        products = db.query(ProductModel).options(
            selectinload(ProductModel.part_type),
        ).filter(
            ProductModel.organization_id == org_id,
            ProductModel.id.in_(product_ids),
        ).all()
        by_id = {p.id: p for p in products}
        ordered_products = [by_id.get(pid) for pid in product_ids if by_id.get(pid)]
        if not ordered_products:
            _set_job_state(db, job_id, status="failed", stage="failed", error_summary="Товары не найдены")
            return {"status": "failed", "error": "products not found"}

        _set_job_state(db, job_id, stage="map_rows", total=len(ordered_products))
        product_id_list = [p.id for p in ordered_products]

        # Batch load links and photos to avoid N+1
        links = db.query(ProductAvitoListingLink).filter(
            ProductAvitoListingLink.organization_id == org_id,
            ProductAvitoListingLink.product_id.in_(product_id_list),
        ).all()
        # Используем реальный avito_id (Avito item_id), а не avito_ad_id (internal_code)
        link_map = {l.product_id: (l.avito_id or "") for l in links}

        photos = db.query(ProductPhoto).filter(
            ProductPhoto.organization_id == org_id,
            ProductPhoto.product_id.in_(product_id_list),
        ).all()
        photo_map: dict[int, list[str]] = {}
        for p in photos:
            photo_map.setdefault(p.product_id, []).append(p.photo_url)

        org = db.query(OrganizationModel).filter(OrganizationModel.id == org_id).first()
        org_address = (org.address if org and org.address else "").strip()
        storage_ids = {p.storage_location_id for p in ordered_products if p.storage_location_id}
        storage_rows = (
            db.query(StorageLocationModel).filter(StorageLocationModel.id.in_(list(storage_ids))).all()
            if storage_ids
            else []
        )
        storage_by_id = {s.id: s for s in storage_rows}

        xlsx_path, rel_path = _resolve_saved_autoload_file(org_id)
        existing_bytes = xlsx_path.read_bytes() if xlsx_path.is_file() else None
        export_rows: list[dict[str, Any]] = []
        for idx, product in enumerate(ordered_products, start=1):
            category = _map_avito_category(product)
            storage = storage_by_id.get(product.storage_location_id) if product.storage_location_id else None
            address = ((storage.address if storage and storage.address else "") or org_address).strip()
            raw_photos = photo_map.get(product.id, [])
            photos_for_xlsx = product_photo_urls_for_avito_export(raw_photos)
            description = append_marketplace_site_info(
                product.description or "",
                enabled=bool(getattr(org, "append_marketplace_site_info", False)),
                product=product,
                site_origin=(settings.PUBLIC_BASE_URL or settings.BASE_URL or "").strip(),
            )
            
            export_rows.append(
                {
                    "id": product.id,
                    "internal_code": product.internal_code,
                    "article": product.article,
                    "brand": product.brand,
                    "is_new": product.is_new,
                    "price": product.price,
                    "name": product.name,
                    "description": description,
                    "quantity": product.quantity,
                    "photos": photos_for_xlsx,
                    "avito_id": link_map.get(product.id, ""),
                    "category": "Запчасти и аксессуары",  # Всегда эта категория для листа "Объявления"
                    "template_sheet": "Объявления",  # Всегда этот лист
                    "address": address,
                    "part_type_name": product.part_type.name if product.part_type else "",
                    # NEW: Map to new format fields
                    "availability": "В наличии" if product.quantity > 0 else "Под заказ",
                    "originality": "Оригинал" if product.is_new else "Неоригинал",
                }
            )
            if idx % 100 == 0 or idx == len(ordered_products):
                _set_job_state(db, job_id, processed=idx)

        _set_job_state(db, job_id, stage="upsert_xlsx")
        merged_bytes = upsert_products_to_avito_autoload(
            existing_bytes,
            export_rows,
            public_base_url=(settings.PUBLIC_BASE_URL or settings.BASE_URL or "").strip(),
        )
        xlsx_path.write_bytes(merged_bytes)

        _set_job_state(db, job_id, stage="validate")
        parsed = parse_and_validate_avito_autoload(merged_bytes)

        avito_upload = None
        avito_upload_status = None
        avito_report = None
        avito_token_error = None

        if publish_after_export:
            _set_job_state(db, job_id, stage="publish")
            row = db.query(OrganizationAvitoIntegration).filter(
                OrganizationAvitoIntegration.organization_id == org_id
            ).first()
            if row:
                try:
                    sec = decrypt_secret(row.client_secret_encrypted)
                    token = asyncio.run(avito_api_svc.fetch_access_token(row.client_id, sec))
                    avito_upload_status, avito_upload = asyncio.run(
                        avito_api_svc.upload_autoload_xlsx(token, xlsx_path.name, merged_bytes)
                    )
                    if avito_upload_status in (200, 201):
                        avito_report = asyncio.run(avito_api_svc.get_last_completed_report_v3(token))
                        if avito_report is None:
                            avito_report = asyncio.run(
                                avito_api_svc.get_last_report_v1(token, int(row.avito_user_id))
                            )
                        
                        # After successful publish, re-parse xlsx to get AvitoIds and sync links
                        try:
                            updated_bytes = xlsx_path.read_bytes()
                            updated_parsed = parse_and_validate_avito_autoload(updated_bytes)
                            
                            # Sync AvitoIds to database
                            sync_stats = {"updated": 0, "created": 0, "skipped": 0}
                            for item in updated_parsed.items:
                                internal_code = item.get('unique_ad_id')
                                avito_id_from_xlsx = item.get('avito_id')
                                
                                if not internal_code or not avito_id_from_xlsx:
                                    sync_stats["skipped"] += 1
                                    continue
                                
                                link = db.query(ProductAvitoListingLink).filter(
                                    ProductAvitoListingLink.organization_id == org_id,
                                    ProductAvitoListingLink.avito_ad_id == str(internal_code),
                                ).first()
                                
                                if link:
                                    if link.avito_id != str(avito_id_from_xlsx):
                                        link.avito_id = str(avito_id_from_xlsx)
                                        sync_stats["updated"] += 1
                                else:
                                    product = db.query(ProductModel).filter(
                                        ProductModel.organization_id == org_id,
                                        ProductModel.internal_code == str(internal_code),
                                    ).first()
                                    
                                    if product:
                                        new_link = ProductAvitoListingLink(
                                            organization_id=org_id,
                                            product_id=product.id,
                                            avito_ad_id=str(internal_code),
                                            avito_id=str(avito_id_from_xlsx),
                                        )
                                        db.add(new_link)
                                        sync_stats["created"] += 1
                                    else:
                                        sync_stats["skipped"] += 1
                            
                            db.commit()
                            print(f"✅ Avito ID sync after publish: {sync_stats}")
                            
                            # Update parsed items with fresh data
                            parsed = updated_parsed
                        except Exception as sync_err:
                            print(f"⚠️ Failed to sync Avito IDs after publish: {sync_err}")
                            db.rollback()
                except Exception as e:
                    avito_token_error = str(e)

        _save_autoload_cache(
            db,
            org_id,
            saved_path=rel_path,
            items=parsed.items,
            local_validation_ok=parsed.local_ok,
            local_errors=parsed.local_errors,
            sheets_parsed=parsed.sheets_parsed,
            avito_upload=avito_upload,
            avito_upload_status=avito_upload_status,
            avito_report=avito_report,
            avito_token_error=avito_token_error,
        )

        result = {
            "saved_path": rel_path,
            "exported_count": len(export_rows),
            "local_validation_ok": parsed.local_ok,
            "local_errors_count": len(parsed.local_errors),
            "unmapped_count": 0,  # Все товары имеют категорию "Запчасти и аксессуары"
            "avito_upload_status": avito_upload_status,
        }
        _set_job_state(
            db,
            job_id,
            status="completed",
            stage="completed",
            processed=len(export_rows),
            total=len(export_rows),
            result_file_ref=rel_path,
            result=result,
        )
        return {"status": "completed", **result}
    except Exception as e:
        _set_job_state(db, job_id, status="failed", stage="failed", error_summary=str(e))
        raise
    finally:
        db.close()


@celery_app.task(bind=True, max_retries=3)
def run_avito_publish_job(self, job_id: int):
    db = SessionLocal()
    try:
        job = db.query(AvitoAutoloadJob).filter(AvitoAutoloadJob.id == job_id).first()
        if not job:
            return {"status": "failed", "error": "job not found"}
        org_id = job.organization_id
        if not is_avito_pro_active(db, org_id):
            _set_job_state(
                db,
                job_id,
                status="failed",
                stage="failed",
                error_summary="Подписка Avito Pro истекла или нет доступа к API Avito",
            )
            return {"status": "failed", "error": "avito pro inactive"}
        _set_job_state(db, job_id, status="processing", stage="load_file")

        xlsx_path, rel_path = _resolve_saved_path_from_cache(db, org_id)
        if not xlsx_path:
            _set_job_state(db, job_id, status="failed", stage="failed", error_summary="Файл автозагрузки не найден")
            return {"status": "failed", "error": "autoload file not found"}

        row = db.query(OrganizationAvitoIntegration).filter(
            OrganizationAvitoIntegration.organization_id == org_id
        ).first()
        if not row:
            _set_job_state(db, job_id, status="failed", stage="failed", error_summary="Интеграция Авито не настроена")
            return {"status": "failed", "error": "integration not configured"}

        _set_job_state(db, job_id, stage="publish")
        file_bytes = xlsx_path.read_bytes()
        avito_upload = None
        avito_upload_status = None
        avito_report = None
        avito_token_error = None
        try:
            sec = decrypt_secret(row.client_secret_encrypted)
            token = asyncio.run(avito_api_svc.fetch_access_token(row.client_id, sec))
            avito_upload_status, avito_upload = asyncio.run(
                avito_api_svc.upload_autoload_xlsx(token, xlsx_path.name, file_bytes)
            )
            if avito_upload_status in (200, 201):
                avito_report = asyncio.run(avito_api_svc.get_last_completed_report_v3(token))
                if avito_report is None:
                    avito_report = asyncio.run(avito_api_svc.get_last_report_v1(token, int(row.avito_user_id)))
                
                # After successful publish, re-parse xlsx to get AvitoIds and sync links
                try:
                    updated_bytes = xlsx_path.read_bytes()
                    updated_parsed = parse_and_validate_avito_autoload(updated_bytes)
                    
                    # Sync AvitoIds to database
                    sync_stats = {"updated": 0, "created": 0, "skipped": 0}
                    for item in updated_parsed.items:
                        internal_code = item.get('unique_ad_id')
                        avito_id_from_xlsx = item.get('avito_id')
                        
                        if not internal_code or not avito_id_from_xlsx:
                            sync_stats["skipped"] += 1
                            continue
                        
                        link = db.query(ProductAvitoListingLink).filter(
                            ProductAvitoListingLink.organization_id == org_id,
                            ProductAvitoListingLink.avito_ad_id == str(internal_code),
                        ).first()
                        
                        if link:
                            if link.avito_id != str(avito_id_from_xlsx):
                                link.avito_id = str(avito_id_from_xlsx)
                                sync_stats["updated"] += 1
                        else:
                            product = db.query(ProductModel).filter(
                                ProductModel.organization_id == org_id,
                                ProductModel.internal_code == str(internal_code),
                            ).first()
                            
                            if product:
                                new_link = ProductAvitoListingLink(
                                    organization_id=org_id,
                                    product_id=product.id,
                                    avito_ad_id=str(internal_code),
                                    avito_id=str(avito_id_from_xlsx),
                                )
                                db.add(new_link)
                                sync_stats["created"] += 1
                            else:
                                sync_stats["skipped"] += 1
                    
                    db.commit()
                    print(f"✅ Avito ID sync after publish job: {sync_stats}")
                except Exception as sync_err:
                    print(f"⚠️ Failed to sync Avito IDs after publish job: {sync_err}")
                    db.rollback()
        except Exception as e:
            avito_token_error = str(e)

        cache = db.query(OrganizationAvitoAutoloadCache).filter(
            OrganizationAvitoAutoloadCache.organization_id == org_id
        ).first()
        _save_autoload_cache(
            db,
            org_id,
            saved_path=rel_path,
            items=_json_loads(cache.items_json if cache else None, []),
            local_validation_ok=bool(cache.local_validation_ok) if cache else True,
            local_errors=_json_loads(cache.local_errors_json if cache else None, []),
            sheets_parsed=_json_loads(cache.sheets_parsed_json if cache else None, []),
            avito_upload=avito_upload,
            avito_upload_status=avito_upload_status,
            avito_report=avito_report,
            avito_token_error=avito_token_error,
        )

        result = {
            "saved_path": rel_path,
            "avito_upload_status": avito_upload_status,
            "avito_report": avito_report,
            "avito_token_error": avito_token_error,
        }
        _set_job_state(
            db,
            job_id,
            status="completed" if avito_upload_status in (200, 201) else "failed",
            stage="completed" if avito_upload_status in (200, 201) else "failed",
            result_file_ref=rel_path,
            result=result,
            error_summary=avito_token_error if avito_upload_status not in (200, 201) else None,
        )
        return result
    except Exception as e:
        _set_job_state(db, job_id, status="failed", stage="failed", error_summary=str(e))
        raise
    finally:
        db.close()
