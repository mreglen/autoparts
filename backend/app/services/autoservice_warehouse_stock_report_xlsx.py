from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook

from app.services.autoservice_warehouse_stock_report import (
    WarehouseStockReportFilters,
    build_warehouse_stock_report,
)
from app.services.finance_xlsx_export import HEADER_FONT, _autosize_columns, _write_header

UNIT_LABELS = {
    "pcs": "шт.",
    "l": "л",
    "kg": "кг",
}


def _month_label(year: int, month: int) -> str:
    return f"{month:02d}.{year}"


def _sheet_summary(wb: Workbook, report: dict, filters: WarehouseStockReportFilters) -> None:
    ws = wb.active
    ws.title = "Сводка"
    ws["A1"] = "Остатки на складе автосервиса"
    ws["A1"].font = HEADER_FONT
    ws["A2"] = "Месяц"
    ws["B2"] = _month_label(report["year"], report["month"])
    ws["A3"] = "Остаток на дату"
    ws["B3"] = str(report["as_of"])
    if filters.q:
        ws["A4"] = "Поиск"
        ws["B4"] = filters.q
    ws["A5"] = "Скрыть нулевые"
    ws["B5"] = "Да" if filters.hide_zero else "Нет"

    summary = report["summary"]
    row = 7
    metrics = [
        ("Позиций с остатком", summary["positions"]),
        ("Сумма остатков на конец, ₽", summary["closing_value"]),
        ("Сумма остатков на начало, ₽", summary["opening_value"]),
        ("Приход за месяц, шт.", summary["received_qty"]),
        ("Расход за месяц, шт.", summary["expensed_qty"]),
        ("Строк в отчёте", len(report["items"])),
    ]
    for label, value in metrics:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        row += 1
    _autosize_columns(ws)


def _sheet_items(wb: Workbook, report: dict) -> None:
    ws = wb.create_sheet("Остатки")
    is_current = report["is_current_month"]
    headers = [
        "Бренд",
        "Артикул",
        "Наименование",
        "Ед.",
        "Цена, ₽",
        "Остаток на начало",
        "Приход",
        "Расход",
        "Остаток на конец",
    ]
    if is_current:
        headers.extend(["Резерв", "Доступно"])
    headers.append("Сумма, ₽")
    _write_header(ws, headers)

    for index, row in enumerate(report["items"], start=2):
        ws.cell(row=index, column=1, value=row["brand"])
        ws.cell(row=index, column=2, value=row["article"])
        ws.cell(row=index, column=3, value=row["name"])
        ws.cell(row=index, column=4, value=UNIT_LABELS.get(row["unit"], row["unit"]))
        ws.cell(row=index, column=5, value=row["unit_price"])
        ws.cell(row=index, column=6, value=row["opening_qty"])
        ws.cell(row=index, column=7, value=row["received_qty"])
        ws.cell(row=index, column=8, value=row["expensed_qty"])
        ws.cell(row=index, column=9, value=row["closing_qty"])
        col = 10
        if is_current:
            ws.cell(row=index, column=col, value=row["reserved_qty"])
            col += 1
            ws.cell(row=index, column=col, value=row["available_qty"])
            col += 1
        ws.cell(row=index, column=col, value=row["stock_amount"])
    _autosize_columns(ws)


def build_warehouse_stock_workbook_bytes(db, org_id: str, filters: WarehouseStockReportFilters) -> bytes:
    report = build_warehouse_stock_report(db, org_id, filters)
    wb = Workbook()
    _sheet_summary(wb, report, filters)
    _sheet_items(wb, report)
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
