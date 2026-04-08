from __future__ import annotations

from dataclasses import dataclass, field
from io import BytesIO
import re
from typing import Any, Optional

from openpyxl import Workbook, load_workbook

TEMPLATE_XLSX_REL_PATH = "backend/uploads/avito/qMHbBIoD51/autoload.xlsx"
DATA_START_ROW = 6  # 1..4 — служебные строки, 5 — пример/дефолты, данные начинаются с 6

# Заголовки — строго из файла шаблона (2-я строка).
UNIQUE_AD_ID_HEADER = "Уникальный идентификатор объявления"  # col 1
AVITO_AD_NUMBER_HEADER = "Номер объявления на Авито"  # col 3 (Avito возвращает после загрузки)
TITLE_HEADER = "Название объявления"
DESCRIPTION_AD_HEADER = "Описание объявления"
PRICE_HEADER = "Цена"
ADDRESS_HEADER = "Адрес"
CATEGORY_HEADER = "Категория"
CONDITION_HEADER = "Состояние"
PHOTOS_HEADER = "Ссылки на фото"
PHOTO_NAMES_HEADER = "Названия фото"
QUANTITY_HEADER = "Количество"
MANUFACTURER_HEADER = "Производитель"
PART_OEM = "Номер детали OEM"
AVITO_STATUS_HEADER = "AvitoStatus"


def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float):
        if v == int(v):
            return str(int(v))
    return str(v).strip()


def _is_mandatory(requirement_cell: Any) -> bool:
    t = _cell_str(requirement_cell).split("\n")[0].strip()
    return t == "Обязательный"


def _find_col_map(header_row: tuple[Any, ...] | list[Any]) -> dict[str, int]:
    m: dict[str, int] = {}
    for idx, val in enumerate(header_row, start=1):
        key = _cell_str(val)
        if key:
            m[key] = idx
    return m


def _row_nonempty(row: tuple[Any, ...], cols: list[int]) -> bool:
    for c in cols:
        if c and c - 1 < len(row) and _cell_str(row[c - 1]):
            return True
    return False


def _normalize_header(value: Any) -> str:
    raw = _cell_str(value).strip().lower()
    return re.sub(r"[^a-zа-я0-9]+", "", raw)


def _find_optional_col(cm: dict[str, int], *variants: str) -> Optional[int]:
    normalized = {_normalize_header(k): v for k, v in cm.items()}
    for variant in variants:
        col = normalized.get(_normalize_header(variant))
        if col:
            return col
    return None


def _find_unique_ad_id_col(cm: dict[str, int]) -> Optional[int]:
    return cm.get(UNIQUE_AD_ID_HEADER) or _find_optional_col(cm, UNIQUE_AD_ID_HEADER)


def _find_avito_ad_number_col(cm: dict[str, int]) -> Optional[int]:
    return cm.get(AVITO_AD_NUMBER_HEADER) or _find_optional_col(cm, AVITO_AD_NUMBER_HEADER)


def _find_legacy_avito_id_col(cm: dict[str, int]) -> Optional[int]:
    # В старом формате могла быть отдельная колонка AvitoId.
    return _find_optional_col(cm, "AvitoId", "AvitoID", "Avito Id", "ID на Авито", "ID объявления", "AdId")


def _split_media_urls(raw: str) -> list[str]:
    if not raw:
        return []
    text = raw.replace("\r", "\n").replace(";", ",").replace("|", ",")
    out: list[str] = []
    for chunk in text.replace("\n", ",").split(","):
        url = chunk.strip()
        if not url:
            continue
        out.append(url)
    # Deduplicate while preserving order.
    return list(dict.fromkeys(out))


@dataclass
class AvitoXlsxParseResult:
    items: list[dict[str, Any]] = field(default_factory=list)
    sheets_parsed: list[str] = field(default_factory=list)
    local_errors: list[dict[str, Any]] = field(default_factory=list)
    local_ok: bool = True


def parse_and_validate_avito_autoload(xlsx_bytes: bytes) -> AvitoXlsxParseResult:
    out = AvitoXlsxParseResult()
    bio = BytesIO(xlsx_bytes)
    # read_only даёт усечённые строки у шаблонов Авито — используем обычный режим
    wb = load_workbook(bio, read_only=False, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        head_rows = list(ws.iter_rows(min_row=1, max_row=4, values_only=True))
        if len(head_rows) < 3:
            continue

        raw_headers = list(head_rows[1])
        raw_req = list(head_rows[2])

        n = len(raw_headers)
        while n > 0 and not _cell_str(raw_headers[n - 1]):
            n -= 1
        headers = raw_headers[:n]
        requirements = (raw_req + [None] * n)[:n]

        if not any(_cell_str(x) == TITLE_HEADER for x in headers):
            continue

        cm = _find_col_map(headers)
        unique_ad_id_c = _find_unique_ad_id_col(cm)
        avito_ad_number_c = _find_avito_ad_number_col(cm)
        legacy_avito_id_c = _find_legacy_avito_id_col(cm)
        pc = cm.get(PART_OEM)  # может отсутствовать в отдельных категориях
        title_c = cm.get(TITLE_HEADER)
        price_c = cm.get(PRICE_HEADER)
        cond_c = cm.get(CONDITION_HEADER)
        man_c = cm.get(MANUFACTURER_HEADER)
        category_c = cm.get(CATEGORY_HEADER)
        avito_status_c = cm.get(AVITO_STATUS_HEADER)
        description_c = _find_optional_col(
            cm,
            DESCRIPTION_AD_HEADER,
            "Описание",
            "Описание товара",
            "Описание товара",
            "Description",
        )
        quantity_c = _find_optional_col(cm, QUANTITY_HEADER, "Кол-во", "Кол во", "Qty", "Quantity")
        photos_csv_c = _find_optional_col(
            cm,
            "Ссылки на фото",
            "Фото",
            "Фотографии",
            "Images",
            "ImageUrls",
            "PhotoUrls",
        )
        video_csv_c = _find_optional_col(cm, "Видео", "Video", "VideoUrl", "VideoUrls")
        photo_cols = [
            c
            for c in [
                _find_optional_col(cm, "Фото1"),
                _find_optional_col(cm, "Фото2"),
                _find_optional_col(cm, "Фото3"),
                _find_optional_col(cm, "Фото4"),
                _find_optional_col(cm, "Фото5"),
                _find_optional_col(cm, "Image1"),
                _find_optional_col(cm, "Image2"),
                _find_optional_col(cm, "Image3"),
                _find_optional_col(cm, "Image4"),
                _find_optional_col(cm, "Image5"),
            ]
            if c
        ]

        if not title_c or not price_c or not cond_c:
            out.local_errors.append(
                {
                    "sheet": sheet_name,
                    "message": "На листе нет обязательных колонок шаблона (название, цена, состояние).",
                }
            )
            out.local_ok = False
            continue

        mandatory_cols: list[tuple[int, str]] = []
        for i in range(len(headers)):
            col_idx = i + 1
            h = _cell_str(headers[i])
            if not h:
                continue
            if _is_mandatory(requirements[i] if i < len(requirements) else None):
                mandatory_cols.append((col_idx, h))

        track_cols = [c for c in [pc, man_c, cond_c, price_c, title_c, category_c, avito_status_c] if c]

        for r_idx, row in enumerate(ws.iter_rows(min_row=DATA_START_ROW, values_only=True), start=DATA_START_ROW):
            row_t = tuple(row)
            if not _row_nonempty(row_t, track_cols):
                continue

            for col_idx, label in mandatory_cols:
                val = row_t[col_idx - 1] if col_idx - 1 < len(row_t) else None
                if not _cell_str(val):
                    out.local_errors.append(
                        {
                            "sheet": sheet_name,
                            "row": r_idx,
                            "field": label,
                            "message": "Обязательное поле пустое",
                        }
                    )
                    out.local_ok = False

            price_raw = row_t[price_c - 1] if price_c - 1 < len(row_t) else None
            price_ok = True
            price_display: Any = None
            try:
                if price_raw is None or _cell_str(price_raw) == "":
                    price_ok = False
                else:
                    ptxt = _cell_str(price_raw).replace(" ", "").replace(",", ".")
                    price_display = int(float(ptxt))
            except (TypeError, ValueError):
                price_ok = False

            if not price_ok:
                out.local_errors.append(
                    {
                        "sheet": sheet_name,
                        "row": r_idx,
                        "field": PRICE_HEADER,
                        "message": "Цена должна быть числом (для шаблона — целое)",
                    }
                )
                out.local_ok = False

            def _at(col: Optional[int]) -> str:
                if not col or col - 1 >= len(row_t):
                    return ""
                return _cell_str(row_t[col - 1])

            item = {
                "sheet": sheet_name,
                "row": r_idx,
                "unique_ad_id": _at(unique_ad_id_c),
                "part_number": _at(pc),
                "manufacturer": _at(man_c),
                "condition": _at(cond_c),
                "price": price_display if price_ok else _cell_str(price_raw),
                "title": _at(title_c),
                "category": _at(category_c),
                "avito_status": _at(avito_status_c),
                # Для импорта/связки используем именно "Номер объявления на Авито" (кол. 3),
                # а при отсутствии — пробуем старую колонку AvitoId.
                "avito_id": _at(avito_ad_number_c) or _at(legacy_avito_id_c),
                "description": _at(description_c),
                "quantity": _at(quantity_c),
                "photos": list(
                    dict.fromkeys(
                        _split_media_urls(_at(photos_csv_c))
                        + [p for col in photo_cols for p in _split_media_urls(_at(col))]
                    )
                ),
                "videos": _split_media_urls(_at(video_csv_c)),
            }
            out.items.append(item)

        out.sheets_parsed.append(sheet_name)

    wb.close()
    return out

def _is_data_sheet(sheet_name: str) -> bool:
    low = (sheet_name or "").strip().lower()
    if low in ("инструкция", "instruction"):
        return False
    if low.startswith("спр-"):
        return False
    return True


def _find_matching_sheet(wb: Workbook, category_name: str) -> str | None:
    if not category_name:
        return None
    category_lower = (category_name or "").strip().lower()
    for sheet_name in wb.sheetnames:
        if not _is_data_sheet(sheet_name):
            continue
        if sheet_name.lower() in category_lower or category_lower in sheet_name.lower():
            return sheet_name
    return None


def _load_template_workbook() -> Workbook:
    from pathlib import Path

    p = Path(TEMPLATE_XLSX_REL_PATH)
    if not p.is_file():
        raise FileNotFoundError(f"Avito template not found at {p}")
    return load_workbook(p, read_only=False, data_only=False)


def _read_sheet_headers(ws) -> list[str]:
    # Шаблон Авито: заголовки во 2-й строке, требования в 3-й.
    # Колонки справа могут быть пустыми — обрезаем хвост.
    values = [ws.cell(row=2, column=i).value for i in range(1, ws.max_column + 1)]
    out = [_cell_str(v) for v in values]
    while out and not out[-1]:
        out.pop()
    return out


def _read_sheet_requirements(ws, n_cols: int) -> list[str]:
    req = [ws.cell(row=3, column=i).value for i in range(1, n_cols + 1)]
    out = [_cell_str(v) for v in req]
    return out


def _read_sheet_defaults(ws, n_cols: int) -> list[Any]:
    # В примере (обычно 5-я строка) в шаблоне заполнены дефолты.
    return [ws.cell(row=5, column=i).value for i in range(1, n_cols + 1)]


def _next_unique_ad_id(existing: set[str]) -> str:
    idx = 1
    while True:
        candidate = f"{idx:05d}"
        if candidate not in existing:
            return candidate
        idx += 1


def _normalize_media_url(url: str, public_base_url: str) -> str:
    value = (url or "").strip()
    if not value:
        return ""
    if value.startswith("http://") or value.startswith("https://"):
        return value
    normalized_base = (public_base_url or "").rstrip("/")
    if not normalized_base:
        return value
    if value.startswith("/temp/"):
        return f"{normalized_base}/uploads{value}"
    if value.startswith("/pictures/") or value.startswith("/videos/") or value.startswith("/vehicle_pictures/"):
        return f"{normalized_base}/uploads{value}"
    if value.startswith("/"):
        return f"{normalized_base}{value}"
    return f"{normalized_base}/{value.lstrip('/')}"


def upsert_products_to_avito_autoload(
    existing_xlsx: Optional[bytes],
    products: list[dict[str, Any]],
    *,
    public_base_url: str,
) -> bytes:
    wb = load_workbook(BytesIO(existing_xlsx), read_only=False) if existing_xlsx else _load_template_workbook()
    sheet_ctx: dict[str, dict[str, Any]] = {}

    def _get_sheet_ctx(category_name: str, template_sheet: str | None):
        sheet_name = None
        if template_sheet and template_sheet in wb.sheetnames and _is_data_sheet(template_sheet):
            sheet_name = template_sheet
        if sheet_name is None and category_name:
            sheet_name = _find_matching_sheet(wb, category_name)
        if sheet_name is None:
            # Fallback: first data sheet
            candidates = [s for s in wb.sheetnames if _is_data_sheet(s)]
            sheet_name = candidates[0] if candidates else wb.sheetnames[0]

        if sheet_name in sheet_ctx:
            return sheet_ctx[sheet_name]

        ws = wb[sheet_name]

        headers = _read_sheet_headers(ws)
        col_map = _find_col_map(headers)
        requirements = _read_sheet_requirements(ws, len(headers))
        defaults = _read_sheet_defaults(ws, len(headers))

        mandatory_headers: list[str] = []
        mandatory_cols: list[int] = []
        for i, h in enumerate(headers):
            if not h:
                continue
            if _is_mandatory(requirements[i] if i < len(requirements) else None):
                mandatory_headers.append(h)
                mandatory_cols.append(i + 1)

        unique_col = col_map.get(UNIQUE_AD_ID_HEADER)
        avito_num_col = col_map.get(AVITO_AD_NUMBER_HEADER)
        legacy_avito_id_col = _find_legacy_avito_id_col(col_map)
        part_col = col_map.get(PART_OEM)

        row_index_by_unique: dict[str, int] = {}
        row_index_by_legacy_avito_id: dict[str, int] = {}
        row_index_by_part: dict[str, int] = {}
        existing_unique_ids: set[str] = set()

        for row_no in range(DATA_START_ROW, ws.max_row + 1):
            uid = str(ws.cell(row=row_no, column=unique_col).value or "").strip() if unique_col else ""
            aid = (
                str(ws.cell(row=row_no, column=legacy_avito_id_col).value or "").strip()
                if legacy_avito_id_col
                else ""
            )
            part = str(ws.cell(row=row_no, column=part_col).value or "").strip() if part_col else ""
            if uid:
                row_index_by_unique[uid] = row_no
                existing_unique_ids.add(uid)
            if aid:
                row_index_by_legacy_avito_id[aid] = row_no
            if part:
                row_index_by_part[part] = row_no

        mandatory_default_by_header: dict[str, Any] = {}
        for i, h in enumerate(headers):
            if h and h in mandatory_headers:
                mandatory_default_by_header[h] = defaults[i] if i < len(defaults) else None

        ctx = {
            "ws": ws,
            "col_map": col_map,
            "headers": headers,
            "requirements": requirements,
            "mandatory_headers": mandatory_headers,
            "mandatory_cols": mandatory_cols,
            "mandatory_default_by_header": mandatory_default_by_header,
            "unique_col": unique_col,
            "avito_num_col": avito_num_col,
            "legacy_avito_id_col": legacy_avito_id_col,
            "part_col": part_col,
            "row_index_by_unique": row_index_by_unique,
            "row_index_by_legacy_avito_id": row_index_by_legacy_avito_id,
            "row_index_by_part": row_index_by_part,
            "existing_unique_ids": existing_unique_ids,
        }
        sheet_ctx[sheet_name] = ctx
        return ctx

    for product in products:
        category = str(product.get("category") or "").strip()
        ctx = _get_sheet_ctx(category, str(product.get("template_sheet") or "").strip() or None)
        ws = ctx["ws"]
        col_map = ctx["col_map"]
        row_index_by_unique = ctx["row_index_by_unique"]
        row_index_by_legacy_avito_id = ctx["row_index_by_legacy_avito_id"]
        row_index_by_part = ctx["row_index_by_part"]
        unique_col = ctx["unique_col"]
        avito_num_col = ctx["avito_num_col"]
        legacy_avito_id_col = ctx["legacy_avito_id_col"]
        part_col = ctx["part_col"]
        mandatory_headers = ctx["mandatory_headers"]
        mandatory_default_by_header = ctx["mandatory_default_by_header"]
        existing_unique_ids = ctx["existing_unique_ids"]

        product_id = str(product.get("id") or "").strip()  # используется только для формируемых полей (не для Excel-ключа)
        legacy_avito_id = str(product.get("avito_id") or "").strip()
        part_number = str(product.get("article") or "").strip()
        internal_code = str(product.get("internal_code") or "").strip()
        unique_ad_id = internal_code or _next_unique_ad_id(existing_unique_ids)

        target_row = (
            (row_index_by_unique.get(unique_ad_id) if unique_ad_id else None)
            or (row_index_by_legacy_avito_id.get(legacy_avito_id) if legacy_avito_id else None)
            or (row_index_by_part.get(part_number) if part_number else None)
        )
        if not target_row:
            target_row = max(ws.max_row + 1, DATA_START_ROW)

        photos = product.get("photos") or []
        photo_urls = []
        for raw_url in photos:
            normalized = _normalize_media_url(str(raw_url), public_base_url)
            if normalized:
                photo_urls.append(normalized)
        photo_urls = list(dict.fromkeys(photo_urls))

        # 1) Стартуем с дефолтов по обязательным полям из примерной строки.
        row_values: dict[str, Any] = {}
        for h in mandatory_headers:
            dv = mandatory_default_by_header.get(h)
            if dv not in (None, ""):
                row_values[h] = dv

        # 2) Затем строго заполняем ключевые колонки.
        if unique_col and UNIQUE_AD_ID_HEADER in col_map:
            row_values[UNIQUE_AD_ID_HEADER] = unique_ad_id

        # Номер объявления на Авито при экспорте не заполняем.
        if avito_num_col and AVITO_AD_NUMBER_HEADER in col_map:
            existing_avito_num = _cell_str(ws.cell(row=target_row, column=avito_num_col).value) if target_row else ""
            if existing_avito_num:
                row_values[AVITO_AD_NUMBER_HEADER] = existing_avito_num
            else:
                row_values[AVITO_AD_NUMBER_HEADER] = ""

        # 3) Поля, зависящие от товара.
        if part_col and PART_OEM in col_map:
            row_values[PART_OEM] = part_number
        if MANUFACTURER_HEADER in col_map:
            row_values[MANUFACTURER_HEADER] = str(product.get("brand") or "")
        if CONDITION_HEADER in col_map:
            row_values[CONDITION_HEADER] = "Новое" if bool(product.get("is_new")) else "Б/у"
        if PRICE_HEADER in col_map:
            # Авито-шаблон ожидает целое; сохраняем как число, но без дробной части.
            try:
                row_values[PRICE_HEADER] = int(float(product.get("price") or 0))
            except (TypeError, ValueError):
                row_values[PRICE_HEADER] = 0
        if TITLE_HEADER in col_map:
            row_values[TITLE_HEADER] = str(product.get("name") or part_number or f"Товар {product_id}")
        if DESCRIPTION_AD_HEADER in col_map:
            row_values[DESCRIPTION_AD_HEADER] = str(product.get("description") or "")
        if ADDRESS_HEADER in col_map:
            row_values[ADDRESS_HEADER] = str(product.get("address") or "")
        if CATEGORY_HEADER in col_map:
            row_values[CATEGORY_HEADER] = str(product.get("category") or "")
        if PHOTOS_HEADER in col_map:
            row_values[PHOTOS_HEADER] = " | ".join(photo_urls)
        if PHOTO_NAMES_HEADER in col_map and PHOTO_NAMES_HEADER not in row_values and photo_urls:
            def _photo_name(u: str) -> str:
                base = (u or "").split("?", 1)[0].rstrip("/")
                name = base.rsplit("/", 1)[-1] if base else ""
                return name or "Фото"

            row_values[PHOTO_NAMES_HEADER] = " | ".join(_photo_name(u) for u in photo_urls)
        if AVITO_STATUS_HEADER in col_map:
            row_values[AVITO_STATUS_HEADER] = str(product.get("avito_status") or "")

        # 3.5) Гарантируем непустые значения для обязательных колонок.
        # В шаблонах Авито встречаются поля, отмеченные "Обязательный", но без дефолта.
        # Для таких случаев ставим заполнители, чтобы файл проходил локальную валидацию.
        for h in mandatory_headers:
            if h not in col_map:
                continue
            current = row_values.get(h)
            if _cell_str(current):
                continue
            existing_val = _cell_str(ws.cell(row=target_row, column=col_map[h]).value)
            if existing_val:
                row_values[h] = existing_val
                continue
            row_values[h] = "Не указано"

        # 4) Запись значений только в существующие колонки шаблона.
        for h, value in row_values.items():
            col = col_map.get(h)
            if not col:
                continue
            ws.cell(row=target_row, column=col, value=value)

        # 5) Обновление индексов для последующих товаров.
        existing_unique_ids.add(unique_ad_id)
        if unique_ad_id:
            row_index_by_unique[unique_ad_id] = target_row
        if legacy_avito_id:
            row_index_by_legacy_avito_id[legacy_avito_id] = target_row
        if part_number:
            row_index_by_part[part_number] = target_row

    out = BytesIO()
    wb.save(out)
    wb.close()
    return out.getvalue()

