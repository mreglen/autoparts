from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook

from app.services.autoservice_rossko_sales_report import RosskoSalesReportFilters, build_rossko_sales_report
from app.services.finance_xlsx_export import HEADER_FONT, _autosize_columns, _write_header


def _sheet_summary(wb: Workbook, report: dict) -> None:
    ws = wb.active
    ws.title = "Сводка"
    ws["A1"] = "Продажи Росско"
    ws["A1"].font = HEADER_FONT
    ws["A2"] = "Период"
    ws["B2"] = f"{report['date_from']} — {report['date_to']}"

    summary = report["summary"]
    row = 4
    metrics = [
        ("Операций", summary["count"]),
        ("Продажа, ₽", summary["sale_total"]),
        ("Закупка Росско, ₽", summary["supplier_total"]),
        ("Эквайринг, ₽", summary["acquiring_fee"]),
        ("Возвраты, ₽", summary["refund_total"]),
        ("Маржа, ₽", summary["margin"]),
        ("Доход сайта (7%), ₽", summary["site_income"]),
        ("Доход организации, ₽", summary["organization_income"]),
        ("Ожидают комиссию", summary["pending_count"]),
    ]
    for label, value in metrics:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=float(value) if value is not None else None)
        row += 1
    _autosize_columns(ws)


def _sheet_operations(wb: Workbook, items: list[dict]) -> None:
    ws = wb.create_sheet("Операции")
    headers = [
        "Дата",
        "Заказ №",
        "Росско №",
        "Покупатель",
        "Оплата",
        "Продажа",
        "Закупка",
        "Эквайринг",
        "Возврат",
        "Маржа",
        "Сайт 7%",
        "Организация",
        "Статус комиссии",
    ]
    _write_header(ws, headers)
    for index, row in enumerate(items, start=2):
        ws.cell(row=index, column=1, value=row.get("operation_at"))
        ws.cell(row=index, column=2, value=row.get("order_id"))
        ws.cell(row=index, column=3, value=row.get("rossko_order_id"))
        ws.cell(row=index, column=4, value=row.get("buyer_name"))
        ws.cell(row=index, column=5, value=row.get("payment_method_label"))
        ws.cell(row=index, column=6, value=float(row.get("sale_total") or 0))
        ws.cell(row=index, column=7, value=float(row.get("supplier_total") or 0))
        ws.cell(
            row=index,
            column=8,
            value=float(row["acquiring_fee"]) if row.get("acquiring_fee") is not None else None,
        )
        ws.cell(row=index, column=9, value=float(row.get("refund_amount") or 0))
        ws.cell(
            row=index,
            column=10,
            value=float(row["margin"]) if row.get("margin") is not None else None,
        )
        ws.cell(
            row=index,
            column=11,
            value=float(row["site_income"]) if row.get("site_income") is not None else None,
        )
        ws.cell(
            row=index,
            column=12,
            value=float(row["organization_income"]) if row.get("organization_income") is not None else None,
        )
        ws.cell(
            row=index,
            column=13,
            value="Ожидается" if row.get("pending_acquiring") else "Рассчитано",
        )
    _autosize_columns(ws)


def _sheet_items(wb: Workbook, items: list[dict]) -> None:
    ws = wb.create_sheet("Позиции")
    headers = [
        "Заказ №",
        "Дата",
        "Бренд",
        "Артикул",
        "Наименование",
        "Кол-во",
        "Продажа",
        "Закупка",
        "Эквайринг",
        "Возврат",
        "Маржа",
        "Сайт 7%",
        "Организация",
    ]
    _write_header(ws, headers)
    row_idx = 2
    for order in items:
        for line in order.get("items") or []:
            ws.cell(row=row_idx, column=1, value=order.get("order_id"))
            ws.cell(row=row_idx, column=2, value=order.get("operation_at"))
            ws.cell(row=row_idx, column=3, value=line.get("brand"))
            ws.cell(row=row_idx, column=4, value=line.get("partnumber"))
            ws.cell(row=row_idx, column=5, value=line.get("name"))
            ws.cell(row=row_idx, column=6, value=line.get("quantity"))
            ws.cell(row=row_idx, column=7, value=float(line.get("sale_total") or 0))
            ws.cell(row=row_idx, column=8, value=float(line.get("supplier_total") or 0))
            ws.cell(
                row=row_idx,
                column=9,
                value=float(line["acquiring_fee"]) if line.get("acquiring_fee") is not None else None,
            )
            ws.cell(row=row_idx, column=10, value=float(line.get("refund_amount") or 0))
            ws.cell(
                row=row_idx,
                column=11,
                value=float(line["margin"]) if line.get("margin") is not None else None,
            )
            ws.cell(
                row=row_idx,
                column=12,
                value=float(line["site_income"]) if line.get("site_income") is not None else None,
            )
            ws.cell(
                row=row_idx,
                column=13,
                value=float(line["organization_income"]) if line.get("organization_income") is not None else None,
            )
            row_idx += 1
    _autosize_columns(ws)


def build_rossko_sales_workbook_bytes(db, org_id: str, filters: RosskoSalesReportFilters) -> bytes:
    report = build_rossko_sales_report(db, org_id, filters)
    wb = Workbook()
    _sheet_summary(wb, report)
    _sheet_operations(wb, report["items"])
    _sheet_items(wb, report["items"])
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
