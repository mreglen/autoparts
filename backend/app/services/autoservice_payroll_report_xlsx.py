from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font

from app.services.autoservice_payroll import compute_org_monthly_payroll
from app.services.finance_xlsx_export import HEADER_FONT, _autosize_columns, _write_header

SUBTOTAL_FONT = Font(bold=True)


def _vehicle_label(vehicle: dict | None) -> str:
    if not vehicle:
        return "—"
    parts = [vehicle.get("make"), vehicle.get("model"), vehicle.get("year")]
    base = " ".join(str(part) for part in parts if part)
    plate = vehicle.get("plate")
    if plate:
        return f"{base} ({plate})" if base else str(plate)
    return base or "—"


def _month_label(year: int, month: int) -> str:
    return f"{month:02d}.{year}"


def _sheet_summary(wb: Workbook, report: dict) -> None:
    ws = wb.active
    ws.title = "Сводка"
    ws["A1"] = "Зарплаты автосервиса"
    ws["A1"].font = HEADER_FONT
    ws["A2"] = "Месяц"
    ws["B2"] = _month_label(report["year"], report["month"])

    order_count = sum(employee["completed_orders"] for employee in report["employees"])
    row = 4
    metrics = [
        ("Сотрудников", len(report["employees"])),
        ("Заказ-нарядов", order_count),
        ("К выплате, ₽", report["total"]),
    ]
    for label, value in metrics:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        row += 1
    _autosize_columns(ws)


def _sheet_employees(wb: Workbook, employees: list[dict]) -> None:
    ws = wb.create_sheet("Сотрудники")
    headers = ["Сотрудник", "Заказ-нарядов", "К выплате, ₽"]
    _write_header(ws, headers)
    for index, row in enumerate(employees, start=2):
        ws.cell(row=index, column=1, value=row["name"])
        ws.cell(row=index, column=2, value=row["completed_orders"])
        ws.cell(row=index, column=3, value=row["total"])
    _autosize_columns(ws)


def _sheet_details(wb: Workbook, employees: list[dict]) -> None:
    ws = wb.create_sheet("По заказ-нарядам")
    headers = ["Сотрудник", "Заказ-наряд", "Автомобиль", "Сумма, ₽"]
    _write_header(ws, headers)
    row_index = 2
    for employee in employees:
        for order in employee.get("orders") or []:
            ws.cell(row=row_index, column=1, value=employee["name"])
            ws.cell(row=row_index, column=2, value=order["order_number"])
            ws.cell(row=row_index, column=3, value=_vehicle_label(order.get("vehicle")))
            ws.cell(row=row_index, column=4, value=order["amount"])
            row_index += 1
    _autosize_columns(ws)


def _sheet_works(wb: Workbook, employees: list[dict]) -> None:
    ws = wb.create_sheet("Работы")
    headers = [
        "Сотрудник",
        "Заказ-наряд",
        "Автомобиль",
        "Работа",
        "Кол-во",
        "Цена, ₽",
        "Сумма работы, ₽",
        "Доля, %",
        "Тип начисления",
        "Начисление, ₽",
    ]
    _write_header(ws, headers)
    row_index = 2

    for employee in employees:
        employee_name = employee["name"]
        employee_total = employee["total"]
        employee_start = row_index
        has_rows = False

        for order in employee.get("orders") or []:
            order_number = order["order_number"]
            vehicle = _vehicle_label(order.get("vehicle"))
            works = order.get("works") or []
            if not works:
                ws.cell(row=row_index, column=1, value=employee_name)
                ws.cell(row=row_index, column=2, value=order_number)
                ws.cell(row=row_index, column=3, value=vehicle)
                ws.cell(row=row_index, column=10, value=order["amount"])
                row_index += 1
                has_rows = True
                continue

            for work in works:
                ws.cell(row=row_index, column=1, value=employee_name)
                ws.cell(row=row_index, column=2, value=order_number)
                ws.cell(row=row_index, column=3, value=vehicle)
                ws.cell(row=row_index, column=4, value=work["title"])
                ws.cell(row=row_index, column=5, value=work["qty"])
                ws.cell(row=row_index, column=6, value=work["unit_price"])
                ws.cell(row=row_index, column=7, value=work["line_total"])
                ws.cell(row=row_index, column=8, value=work["percent"])
                ws.cell(row=row_index, column=9, value=work["accrual_type_label"])
                ws.cell(row=row_index, column=10, value=work["amount"])
                row_index += 1
                has_rows = True

        if not has_rows:
            continue

        ws.cell(row=row_index, column=1, value=f"Итого: {employee_name}")
        ws.cell(row=row_index, column=10, value=employee_total)
        for col in (1, 10):
            ws.cell(row=row_index, column=col).font = SUBTOTAL_FONT
        for row_num in range(employee_start, row_index):
            ws.row_dimensions[row_num].outlineLevel = 1
        row_index += 1

    ws.sheet_properties.outlinePr.summaryBelow = True
    _autosize_columns(ws)


def build_payroll_report_workbook_bytes(db, org_id: str, year: int, month: int) -> bytes:
    report = compute_org_monthly_payroll(db, org_id, year, month)
    wb = Workbook()
    _sheet_summary(wb, report)
    _sheet_employees(wb, report["employees"])
    _sheet_details(wb, report["employees"])
    _sheet_works(wb, report["employees"])
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
