from __future__ import annotations

from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook

from app.services.avito_media import normalize_for_xlsx

# Шаблон Drom находится в app/templates/drom/auto-parts-GT.xlsx
TEMPLATE_XLSX_REL_PATH = "app/templates/drom/auto-parts-GT.xlsx"

# Заголовки из шаблона Drom (первая строка)
ARTICLE_HEADER = "Артикул"
NAME_HEADER = "Наименование товара"
CONDITION_HEADER = "Новый/б.у."
MARK_HEADER = "Марка"
MODEL_HEADER = "Модель"
BODY_HEADER = "Кузов"
NUMBER_HEADER = "Номер"
MANUFACTURER_HEADER = "Производитель"
ENGINE_HEADER = "Двигатель"
YEAR_HEADER = "Год"
LR_HEADER = "LR"
FR_HEADER = "FR"
UD_HEADER = "UD"
COLOR_HEADER = "Цвет"
REPLACEMENT_NUMBERS_HEADER = "Номера замен"
NOTE_HEADER = "Примеч."
QUANTITY_HEADER = "Кол-во"
PRICE_HEADER = "Цена"
AVAILABILITY_HEADER = "Наличие"
DELIVERY_TIME_HEADER = "Сроки доставки"
PHOTO_HEADER = "Фотография"
SUPPLIER_HEADER = "Поставщик"
SUPPLIER_INN_HEADER = "ИНН поставщика"
WAREHOUSE_ADDRESS_HEADER = "Адрес склада"

DATA_WRITE_START_ROW = 2


def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float):
        if v == int(v):
            return str(int(v))
    return str(v).strip()


def _find_col_map(header_row: list[Any]) -> dict[str, int]:
    """Создать маппинг заголовок -> номер колонки (1-based)"""
    m: dict[str, int] = {}
    for idx, val in enumerate(header_row, start=1):
        key = _cell_str(val)
        if key:
            m[key] = idx
    return m


@dataclass
class DromXlsxParseResult:
    items: list[dict[str, Any]] = field(default_factory=list)
    local_errors: list[dict[str, Any]] = field(default_factory=list)
    local_ok: bool = True


def parse_and_validate_drom_autoload(xlsx_bytes: bytes) -> DromXlsxParseResult:
    """Распарсить XLSX файл Drom и валидировать"""
    out = DromXlsxParseResult()
    bio = BytesIO(xlsx_bytes)
    
    try:
        wb = load_workbook(bio, read_only=False, data_only=True)
    except Exception as e:
        out.local_ok = False
        out.local_errors.append({"error": f"Не удалось открыть файл: {str(e)}"})
        return out

    if len(wb.sheetnames) == 0:
        out.local_ok = False
        out.local_errors.append({"error": "Файл не содержит листов"})
        return out

    sheet_name = wb.sheetnames[0]  # Sheet0
    ws = wb[sheet_name]
    
    # Читаем первую строку (заголовки)
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    col_map = _find_col_map(header_row)
    
    # Проверяем обязательные поля
    required_headers = [ARTICLE_HEADER, NAME_HEADER, PRICE_HEADER]
    missing_headers = [h for h in required_headers if h not in col_map]
    
    if missing_headers:
        out.local_ok = False
        out.local_errors.append({
            "error": f"Отсутствуют обязательные колонки: {', '.join(missing_headers)}"
        })
        return out
    
    # Читаем данные начиная со второй строки
    article_col = col_map.get(ARTICLE_HEADER)
    name_col = col_map.get(NAME_HEADER)
    price_col = col_map.get(PRICE_HEADER)
    
    if not article_col or not name_col or not price_col:
        out.local_ok = False
        out.local_errors.append({"error": "Не найдены обязательные колонки"})
        return out
    
    max_row = ws.max_row or 0
    for row_idx in range(DATA_WRITE_START_ROW, max_row + 1):
        row_data = {}
        row_values = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
        
        # Проверяем, есть ли данные в строке
        article = _cell_str(row_values[article_col - 1] if article_col <= len(row_values) else None)
        if not article:
            continue  # Пропускаем пустые строки
        
        # Извлекаем все поля
        for header, col_idx in col_map.items():
            if col_idx <= len(row_values):
                row_data[header] = row_values[col_idx - 1]
            else:
                row_data[header] = None
        
        # Валидация
        row_errors = []
        name = _cell_str(row_data.get(NAME_HEADER))
        price = row_data.get(PRICE_HEADER)
        
        if not name:
            row_errors.append("Отсутствует наименование")
        
        try:
            price_val = float(price) if price not in (None, "") else None
            if price_val is None or price_val <= 0:
                row_errors.append("Некорректная цена")
        except (TypeError, ValueError):
            row_errors.append("Некорректная цена")
        
        if row_errors:
            out.local_errors.append({
                "row": row_idx,
                "article": article,
                "errors": row_errors
            })
            out.local_ok = False
        else:
            out.items.append({
                "row": row_idx,
                "article": article,
                "name": name,
                "price": price,
                "quantity": row_data.get(QUANTITY_HEADER),
                "availability": row_data.get(AVAILABILITY_HEADER),
                "photo": row_data.get(PHOTO_HEADER),
                **row_data
            })
    
    wb.close()
    return out


def upsert_products_to_drom_autoload(
    existing_bytes: Optional[bytes],
    export_rows: list[dict[str, Any]],
    public_base_url: str = ""
) -> bytes:
    """Добавить/обновить товары в XLSX файле Drom"""
    
    # Загружаем существующий файл или создаем из шаблона
    if existing_bytes:
        wb = load_workbook(BytesIO(existing_bytes), read_only=False)
    else:
        # Загружаем шаблон
        template_path = Path(__file__).resolve().parents[2] / TEMPLATE_XLSX_REL_PATH
        if not template_path.is_file():
            raise FileNotFoundError(f"Шаблон Drom не найден: {template_path}")
        wb = load_workbook(str(template_path), read_only=False)
    
    ws = wb.active
    
    # Получаем маппинг колонок из заголовков
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    col_map = _find_col_map(header_row)
    
    # Читаем существующие данные для upsert логики
    existing_articles: dict[str, int] = {}  # article -> row_number
    max_row = ws.max_row or 1
    
    for row_idx in range(DATA_WRITE_START_ROW, max_row + 1):
        row_values = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
        article_col = col_map.get(ARTICLE_HEADER)
        if article_col and article_col <= len(row_values):
            article = _cell_str(row_values[article_col - 1])
            if article:
                existing_articles[article] = row_idx
    
    # Добавляем/обновляем товары
    next_row = max_row + 1
    
    for export_row in export_rows:
        article = str(export_row.get("article", "")).strip()
        if not article:
            continue
        
        # Определяем строку для записи
        if article in existing_articles:
            row_idx = existing_articles[article]
        else:
            row_idx = next_row
            existing_articles[article] = row_idx
            next_row += 1
        
        # Записываем данные
        # Артикул
        if ARTICLE_HEADER in col_map:
            ws.cell(row=row_idx, column=col_map[ARTICLE_HEADER], value=article)
        
        # Наименование товара
        if NAME_HEADER in col_map:
            ws.cell(row=row_idx, column=col_map[NAME_HEADER], value=export_row.get("name", ""))
        
        # Новый/б.у.
        if CONDITION_HEADER in col_map:
            is_new = export_row.get("is_new", False)
            ws.cell(row=row_idx, column=col_map[CONDITION_HEADER], value="Новый" if is_new else "б.у.")
        
        # Производитель
        if MANUFACTURER_HEADER in col_map:
            ws.cell(row=row_idx, column=col_map[MANUFACTURER_HEADER], value=export_row.get("brand", ""))
        
        # Номер (дублируем артикул)
        if NUMBER_HEADER in col_map:
            ws.cell(row=row_idx, column=col_map[NUMBER_HEADER], value=article)
        
        # Кол-во
        if QUANTITY_HEADER in col_map:
            ws.cell(row=row_idx, column=col_map[QUANTITY_HEADER], value=export_row.get("quantity", 0))
        
        # Цена
        if PRICE_HEADER in col_map:
            ws.cell(row=row_idx, column=col_map[PRICE_HEADER], value=export_row.get("price", 0))
        
        # Наличие
        if AVAILABILITY_HEADER in col_map:
            quantity = export_row.get("quantity", 0)
            ws.cell(row=row_idx, column=col_map[AVAILABILITY_HEADER], 
                   value="В наличии" if quantity > 0 else "Под заказ")
        
        # Фотография (первое фото)
        if PHOTO_HEADER in col_map:
            photos = export_row.get("photos", [])
            if photos:
                # Нормализуем URL фотографии (добавляем домен для локальных путей)
                photo_url = normalize_for_xlsx(photos[0])
                ws.cell(row=row_idx, column=col_map[PHOTO_HEADER], value=photo_url)
            else:
                ws.cell(row=row_idx, column=col_map[PHOTO_HEADER], value="")
        
        # Адрес склада
        if WAREHOUSE_ADDRESS_HEADER in col_map:
            ws.cell(row=row_idx, column=col_map[WAREHOUSE_ADDRESS_HEADER], 
                   value=export_row.get("storage_address", ""))
    
    # Сохраняем в bytes
    out = BytesIO()
    wb.save(out)
    wb.close()
    
    return out.getvalue()
