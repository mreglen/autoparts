from __future__ import annotations

from datetime import date
from io import BytesIO

from openpyxl import Workbook

from app.services.autoservice_payment_service import list_finance_receipts
from app.services.finance_xlsx_export import HEADER_FONT, _autosize_columns, _write_header

METHOD_LABELS = {
    "card": "Карта",
    "cash": "Наличными",
    "bank": "Расчётный счёт",
}


def _sheet_summary(wb: Workbook, report, date_from: date, date_to: date) -> None:
    ws = wb.active
    ws.title = "Сводка"
    ws["A1"] = "Платежи автосервиса"
    ws["A1"].font = HEADER_FONT
    ws["A2"] = "Период"
    ws["B2"] = f"{date_from.isoformat()} — {date_to.isoformat()}"

    row = 4
    metrics = [
        ("Платежей", report.count),
        ("Сумма, ₽", report.total_amount),
        ("Картой, ₽", report.totals.card),
        ("Наличными, ₽", report.totals.cash),
        ("Расчётный счёт, ₽", report.totals.bank),
    ]
    for label, value in metrics:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        row += 1
    _autosize_columns(ws)


def _sheet_payments(wb: Workbook, items) -> None:
    ws = wb.create_sheet("Платежи")
    headers = [
        "№",
        "Заказ-наряд",
        "Плательщик",
        "Способ",
        "Сумма, ₽",
        "Дата",
    ]
    _write_header(ws, headers)
    for index, row in enumerate(items, start=2):
        ws.cell(row=index, column=1, value=row.sequential_number)
        ws.cell(row=index, column=2, value=row.repair_order_number)
        ws.cell(row=index, column=3, value=row.payer_name)
        ws.cell(row=index, column=4, value=METHOD_LABELS.get(row.method, row.method))
        ws.cell(row=index, column=5, value=row.amount)
        ws.cell(row=index, column=6, value=row.created_at)
    _autosize_columns(ws)


def build_finance_receipts_workbook_bytes(db, org_id: str, date_from: date, date_to: date) -> bytes:
    report = list_finance_receipts(db, org_id=org_id, date_from=date_from, date_to=date_to)
    wb = Workbook()
    _sheet_summary(wb, report, date_from, date_to)
    _sheet_payments(wb, report.items)
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
