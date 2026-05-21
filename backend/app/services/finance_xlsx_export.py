"""
Экспорт финансовых отчётов в XLSX (openpyxl).
"""

from __future__ import annotations

from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.services.finance_reports import (
    CHANNEL_LABELS,
    FinanceFilters,
    build_finance_summary,
    list_finance_inventory,
    list_finance_sales,
    list_finance_stock_ins,
    list_finance_writeoffs,
)

HEADER_FONT = Font(bold=True)


def _write_header(ws, headers: list[str]) -> None:
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = HEADER_FONT


def _autosize_columns(ws, max_width: int = 40) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            val = row[0].value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), max_width)


def _sheet_summary(wb: Workbook, summary: dict) -> None:
    ws = wb.active
    ws.title = "Сводка"
    ws["A1"] = "Финансовый отчёт"
    ws["A1"].font = HEADER_FONT
    ws["A2"] = "Период"
    ws["B2"] = f"{summary['date_from']} — {summary['date_to']}"
    ws["A3"] = "Остатки на дату"
    ws["B3"] = str(summary["as_of_date"])

    row = 5
    metrics = [
        ("Продажи (строк)", summary["sales_count"]),
        ("Продажи (сумма), ₽", summary["sales_total"]),
        ("Списания (строк)", summary["writeoffs_count"]),
        ("Списания (шт.)", summary["writeoffs_qty"]),
        ("Поступления (строк)", summary["stock_in_count"]),
        ("Поступления (шт.)", summary["stock_in_qty"]),
        ("Поступления (сумма), ₽", summary["stock_in_value"]),
        ("Остатки (позиций)", summary["inventory_products"]),
        ("Остатки (шт.)", summary["inventory_qty"]),
        ("Остатки (оценка), ₽", summary["inventory_value"]),
    ]
    for label, value in metrics:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Продажи по каналам").font = HEADER_FONT
    row += 1
    for ch_key, ch_data in (summary.get("sales_by_channel") or {}).items():
        label = ch_data.get("label") or CHANNEL_LABELS.get(ch_key, ch_key)
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=ch_data.get("count", 0))
        ws.cell(row=row, column=3, value=ch_data.get("total", 0))
        row += 1

    row += 1
    ws.cell(row=row, column=1, value=summary.get("inventory_note", ""))
    _autosize_columns(ws)


def _sheet_sales(wb: Workbook, rows: list[dict]) -> None:
    ws = wb.create_sheet("Продажи")
    headers = [
        "Дата",
        "Артикул",
        "Внутр. код",
        "Название",
        "Бренд",
        "Кол-во",
        "Цена",
        "Сумма",
        "Канал",
        "ID Авито",
        "ID заказа Б/У",
        "Причина",
    ]
    _write_header(ws, headers)
    for i, r in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=r.get("movement_date"))
        ws.cell(row=i, column=2, value=r.get("article"))
        ws.cell(row=i, column=3, value=r.get("internal_code"))
        ws.cell(row=i, column=4, value=r.get("name"))
        ws.cell(row=i, column=5, value=r.get("brand"))
        ws.cell(row=i, column=6, value=r.get("quantity"))
        ws.cell(row=i, column=7, value=r.get("unit_price"))
        ws.cell(row=i, column=8, value=r.get("line_total"))
        ws.cell(row=i, column=9, value=r.get("channel_label"))
        ws.cell(row=i, column=10, value=r.get("avito_order_id"))
        ws.cell(row=i, column=11, value=r.get("garage_used_order_item_id"))
        ws.cell(row=i, column=12, value=r.get("reason"))
    _autosize_columns(ws)


def _sheet_writeoffs(wb: Workbook, rows: list[dict]) -> None:
    ws = wb.create_sheet("Списания")
    headers = ["Дата", "Артикул", "Внутр. код", "Название", "Бренд", "Кол-во", "Цена", "Причина"]
    _write_header(ws, headers)
    for i, r in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=r.get("movement_date"))
        ws.cell(row=i, column=2, value=r.get("article"))
        ws.cell(row=i, column=3, value=r.get("internal_code"))
        ws.cell(row=i, column=4, value=r.get("name"))
        ws.cell(row=i, column=5, value=r.get("brand"))
        ws.cell(row=i, column=6, value=r.get("quantity"))
        ws.cell(row=i, column=7, value=r.get("sale_price"))
        ws.cell(row=i, column=8, value=r.get("reason"))
    _autosize_columns(ws)


def _sheet_stock_ins(wb: Workbook, rows: list[dict]) -> None:
    ws = wb.create_sheet("Поступления")
    headers = [
        "Дата",
        "Артикул",
        "Внутр. код",
        "Название",
        "Бренд",
        "Кол-во",
        "Цена",
        "Сумма",
        "Сотрудник",
    ]
    _write_header(ws, headers)
    for i, r in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=r.get("created_at"))
        ws.cell(row=i, column=2, value=r.get("article"))
        ws.cell(row=i, column=3, value=r.get("internal_code"))
        ws.cell(row=i, column=4, value=r.get("name"))
        ws.cell(row=i, column=5, value=r.get("brand"))
        ws.cell(row=i, column=6, value=r.get("quantity"))
        ws.cell(row=i, column=7, value=r.get("unit_price"))
        ws.cell(row=i, column=8, value=r.get("line_total"))
        ws.cell(row=i, column=9, value=r.get("creator_name"))
    _autosize_columns(ws)


def _sheet_inventory(wb: Workbook, rows: list[dict], as_of: date, note: str) -> None:
    ws = wb.create_sheet("Остатки")
    ws["A1"] = f"Остатки на {as_of.isoformat()}"
    ws["A1"].font = HEADER_FONT
    ws["A2"] = note
    headers = [
        "Артикул",
        "Внутр. код",
        "Название",
        "Бренд",
        "Остаток",
        "Цена",
        "Оценка",
        "Приход",
        "Расход",
    ]
    start_row = 4
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col, value=title)
        cell.font = HEADER_FONT
    for i, r in enumerate(rows, start=start_row + 1):
        ws.cell(row=i, column=1, value=r.get("article"))
        ws.cell(row=i, column=2, value=r.get("internal_code"))
        ws.cell(row=i, column=3, value=r.get("name"))
        ws.cell(row=i, column=4, value=r.get("brand"))
        ws.cell(row=i, column=5, value=r.get("quantity"))
        ws.cell(row=i, column=6, value=r.get("unit_price"))
        ws.cell(row=i, column=7, value=r.get("line_total"))
        ws.cell(row=i, column=8, value=r.get("stock_in_qty"))
        ws.cell(row=i, column=9, value=r.get("stock_out_qty"))
    _autosize_columns(ws)


def build_finance_workbook_bytes(
    db: Session,
    organization_id: str,
    filters: FinanceFilters,
) -> bytes:
    summary = build_finance_summary(db, organization_id, filters)
    sales_rows, _ = list_finance_sales(db, organization_id, filters)
    writeoff_rows, _ = list_finance_writeoffs(db, organization_id, filters)
    stock_in_rows, _ = list_finance_stock_ins(db, organization_id, filters)
    inventory_rows, inventory_meta = list_finance_inventory(db, organization_id, filters)

    wb = Workbook()
    _sheet_summary(wb, summary)
    _sheet_sales(wb, sales_rows)
    _sheet_writeoffs(wb, writeoff_rows)
    _sheet_stock_ins(wb, stock_in_rows)
    _sheet_inventory(
        wb,
        inventory_rows,
        filters.as_of_date,
        inventory_meta.get("note", ""),
    )

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
