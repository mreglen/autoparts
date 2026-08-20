from __future__ import annotations

from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.services.autoservice_order_economics import (
    OrderEconomicsFilters,
    build_order_economics_report,
    order_economics_payment_label,
    order_economics_status_label,
)
from app.services.finance_xlsx_export import HEADER_FONT, _autosize_columns, _write_header

STATUS_LABELS = {
    "all": "Все",
    "pending": "Ожидание",
    "in_progress": "В работе",
    "done": "Выполнен",
    "completed": "Закрыт",
    "cancelled": "Отменён",
}

PAYMENT_LABELS = {
    "all": "Все",
    "paid": "Оплачено",
    "partial": "Частично",
    "unpaid": "Долг",
}


def _vehicle_label(vehicle: dict | None) -> str:
    if not vehicle:
        return "—"
    parts = [vehicle.get("make"), vehicle.get("model"), vehicle.get("year")]
    base = " ".join(str(part) for part in parts if part)
    plate = vehicle.get("plate")
    if plate:
        return f"{base} ({plate})" if base else str(plate)
    return base or "—"


def _sheet_summary(wb: Workbook, report: dict, filters: OrderEconomicsFilters) -> None:
    ws = wb.active
    ws.title = "Сводка"
    ws["A1"] = "Экономика заказ-нарядов"
    ws["A1"].font = HEADER_FONT
    ws["A2"] = "Период"
    ws["B2"] = f"{report['date_from']} — {report['date_to']}"
    ws["A3"] = "Статус"
    ws["B3"] = STATUS_LABELS.get(filters.status, filters.status)
    ws["A4"] = "Оплата"
    ws["B4"] = PAYMENT_LABELS.get(filters.payment, filters.payment)
    if filters.q:
        ws["A5"] = "Поиск"
        ws["B5"] = filters.q

    summary = report["summary"]
    row = 7
    metrics = [
        ("Заказ-нарядов", summary["count"]),
        ("Выручка, ₽", summary["revenue"]),
        ("Себестоимость запчастей, ₽", summary["parts_cost"]),
        ("Зарплата, ₽", summary["payroll_total"]),
        ("Чистая прибыль, ₽", summary["net_profit"]),
        ("Оплачено, ₽", summary["paid_amount"]),
        ("Долг, ₽", summary["debt_amount"]),
        ("Неоплаченных заказов", summary["unpaid_count"]),
    ]
    for label, value in metrics:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        row += 1
    _autosize_columns(ws)


def _sheet_orders(wb: Workbook, items: list[dict]) -> None:
    ws = wb.create_sheet("Заказ-наряды")
    headers = [
        "Заказ-наряд",
        "Клиент",
        "Автомобиль",
        "Дата записи",
        "Сумма заказа",
        "Себестоимость запчастей",
        "Зарплата",
        "Чистая прибыль",
        "Оплачено",
        "Долг",
        "Оплата",
        "Статус",
        "Предварительный расчёт",
    ]
    _write_header(ws, headers)
    for index, row in enumerate(items, start=2):
        ws.cell(row=index, column=1, value=row["order_number"])
        ws.cell(row=index, column=2, value=row["client_name"])
        ws.cell(row=index, column=3, value=_vehicle_label(row.get("vehicle")))
        ws.cell(row=index, column=4, value=row["scheduled_at"])
        ws.cell(row=index, column=5, value=row["grand_total"])
        ws.cell(row=index, column=6, value=row["parts_cost"])
        ws.cell(row=index, column=7, value=row["payroll_total"])
        ws.cell(row=index, column=8, value=row["net_profit"])
        ws.cell(row=index, column=9, value=row["paid_amount"])
        ws.cell(row=index, column=10, value=row["remaining_amount"])
        ws.cell(row=index, column=11, value=order_economics_payment_label(row["payment_status"]))
        ws.cell(row=index, column=12, value=order_economics_status_label(row["status"]))
        ws.cell(row=index, column=13, value="Да" if row["is_preliminary"] else "Нет")
    _autosize_columns(ws)


def build_order_economics_workbook_bytes(db, org_id: str, filters: OrderEconomicsFilters) -> bytes:
    report = build_order_economics_report(db, org_id, filters)
    wb = Workbook()
    _sheet_summary(wb, report, filters)
    _sheet_orders(wb, report["items"])
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
