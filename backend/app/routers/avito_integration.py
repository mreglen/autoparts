import json
import logging
import re
import time
from io import BytesIO
from pathlib import Path
from typing import Any, Optional

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.core.auth import get_current_user
from app.core.config import settings
from app.db.database import get_db
from app.models.organization import Organization as OrganizationModel
from app.models.organization_avito_autoload_cache import OrganizationAvitoAutoloadCache
from app.models.organization_avito_integration import OrganizationAvitoIntegration
from app.models.avito_autoload_job import AvitoAutoloadJob
from app.models.product import Product as ProductModel, ProductPhoto, ProductVideo
from app.models.product_avito_listing_link import ProductAvitoListingLink
from app.models.stock_in import StockIn as StockInModel
from app.models.storage_location import StorageLocation as StorageLocationModel
from app.models.user import User as UserModel
from app.schemas.avito_integration import (
    AvitoAutoloadImportRequest,
    AvitoAutoloadImportResponse,
    AvitoAutoloadApplyActionRequest,
    AvitoAutoloadRemoveRowsRequest,
    AvitoAutoloadExportRequest,
    AvitoAutoloadExportAsyncRequest,
    AvitoAutoloadExportResponse,
    AvitoAutoloadJobResponse,
    AvitoAutoloadPublishResponse,
    AvitoAutoloadUploadResponse,
    AvitoAutoloadCategoryTreeResponse,
    AvitoAutoloadSetCategoryRequest,
    AvitoAutoloadSetAdTypeRequest,
    AvitoCredentialsResponse,
    AvitoCredentialsUpdate,
    AvitoLastAutoloadSnapshot,
)
from app.tasks.avito_tasks import run_avito_export_job, run_avito_publish_job, _map_avito_category
from app.services import avito_api as avito_api_svc
from app.services.avito_autoload_xlsx import parse_and_validate_avito_autoload, upsert_products_to_avito_autoload
from app.services.avito_media import ensure_local_pictures, normalize_for_xlsx
from app.utils.avito_crypto import decrypt_secret, encrypt_secret

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/organizations", tags=["Avito"])

SAFE_NAME_RE = re.compile(r"[^A-Za-z0-9._-]+")
ACTION_HEADERS = ("Действие", "Action")
ACTION_VALUES = {
    "publish": "Опубликовать объявление",
    "unpublish": "Снять с публикации",
    "delete": "Удалить объявление",
}

_CATEGORY_TREE_CACHE: dict[str, tuple[float, dict[str, Any]]] = {}
_CATEGORY_TREE_TTL_S = 20 * 60


def _normalize_category_tree(raw: Any) -> list[dict[str, Any]]:
    def to_nodes(obj: Any) -> list[dict[str, Any]]:
        if obj is None:
            return []
        if isinstance(obj, list):
            out: list[dict[str, Any]] = []
            for it in obj:
                out.extend(to_nodes(it))
            return out
        if not isinstance(obj, dict):
            return []

        title = str(
            obj.get("title")
            or obj.get("name")
            or obj.get("label")
            or obj.get("text")
            or obj.get("value")
            or ""
        ).strip()

        children_src = (
            obj.get("children")
            or obj.get("nested")
            or obj.get("items")
            or obj.get("nodes")
            or obj.get("subcategories")
            or []
        )
        children = to_nodes(children_src)
        if not title:
            return children
        return [{"title": title, "children": children}]

    if isinstance(raw, dict):
        for key in ("tree", "categories", "result", "data", "items", "nodes"):
            if key in raw:
                return to_nodes(raw.get(key))
    return to_nodes(raw)


def _remove_existing_avito_xlsx_files(base_dir: Path) -> None:
    """В каталоге организации остаётся не больше одного XLSX — старые удаляем перед новой загрузкой."""
    if not base_dir.is_dir():
        return
    for p in base_dir.iterdir():
        if p.is_file() and p.suffix.lower() == ".xlsx":
            try:
                p.unlink()
            except OSError as e:
                logger.warning("Не удалось удалить старый файл автозагрузки %s: %s", p, e)


def _ensure_org_access(user: UserModel, org_id: str) -> None:
    if user.organization_id != org_id and not user.is_admin:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Нет доступа к организации")


def _org_exists(db: Session, org_id: str) -> OrganizationModel:
    org = db.query(OrganizationModel).filter(OrganizationModel.id == org_id).first()
    if not org:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Организация не найдена")
    return org


def _json_loads(raw: Optional[str], default: Any) -> Any:
    if not raw:
        return default
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        return default


def _get_last_autoload(db: Session, org_id: str) -> Optional[AvitoLastAutoloadSnapshot]:
    cache = (
        db.query(OrganizationAvitoAutoloadCache)
        .filter(OrganizationAvitoAutoloadCache.organization_id == org_id)
        .first()
    )
    if not cache:
        return None

    # Автовосстановление кэша: если файл на диске есть, а items пустые
    # (например, из-за прошлой версии парсера), перепарсим и сохраним.
    try:
        items = _json_loads(cache.items_json, [])
        if cache.saved_path and isinstance(items, list) and len(items) == 0:
            project_root = Path(__file__).resolve().parents[2]
            xlsx_path = project_root / cache.saved_path.lstrip("/")
            if xlsx_path.is_file():
                parsed = parse_and_validate_avito_autoload(xlsx_path.read_bytes())
                cache.items_json = json.dumps(parsed.items, ensure_ascii=False)
                cache.local_validation_ok = bool(parsed.local_ok)
                cache.local_errors_json = json.dumps(parsed.local_errors, ensure_ascii=False)
                cache.sheets_parsed_json = json.dumps(parsed.sheets_parsed, ensure_ascii=False)
                db.commit()
    except Exception:
        db.rollback()

    return AvitoLastAutoloadSnapshot(
        saved_path=cache.saved_path,
        items=_json_loads(cache.items_json, []),
        local_validation_ok=bool(cache.local_validation_ok),
        local_errors=_json_loads(cache.local_errors_json, []),
        sheets_parsed=_json_loads(cache.sheets_parsed_json, []),
        avito_upload=_json_loads(cache.avito_upload_json, None) if cache.avito_upload_json else None,
        avito_upload_status=cache.avito_upload_status,
        avito_report=_json_loads(cache.avito_report_json, None) if cache.avito_report_json else None,
        avito_token_error=cache.avito_token_error,
        updated_at=cache.updated_at.isoformat() if cache.updated_at else None,
        warnings=_json_loads(cache.warnings_json, []) if cache.warnings_json else None,
    )


def _save_autoload_cache(
    db: Session,
    org_id: str,
    *,
    saved_path: str,
    items: list,
    local_validation_ok: bool,
    local_errors: list,
    sheets_parsed: list,
    avito_upload: Any,
    avito_upload_status: Optional[int],
    avito_report: Any,
    avito_token_error: Optional[str],
    warnings: Optional[list] = None,
) -> None:
    row = (
        db.query(OrganizationAvitoAutoloadCache)
        .filter(OrganizationAvitoAutoloadCache.organization_id == org_id)
        .first()
    )
    if not row:
        row = OrganizationAvitoAutoloadCache(organization_id=org_id)
        db.add(row)
    row.saved_path = saved_path
    row.items_json = json.dumps(items, ensure_ascii=False)
    row.local_validation_ok = local_validation_ok
    row.local_errors_json = json.dumps(local_errors, ensure_ascii=False)
    row.sheets_parsed_json = json.dumps(sheets_parsed, ensure_ascii=False)
    row.avito_upload_json = json.dumps(avito_upload, ensure_ascii=False) if avito_upload is not None else None
    row.avito_upload_status = avito_upload_status
    row.avito_report_json = json.dumps(avito_report, ensure_ascii=False) if avito_report is not None else None
    row.avito_token_error = avito_token_error
    row.warnings_json = json.dumps(warnings, ensure_ascii=False) if warnings else None
    db.commit()


def _next_internal_code(db: Session) -> str:
    # В модели Product internal_code уникален глобально (не по organization_id),
    # поэтому подбираем код по всей таблице products.
    # Начинаем с большего номера, чтобы избежать конфликтов
    existing = db.query(ProductModel.internal_code).all()
    existing_codes = {r[0] for r in existing if r[0]}
    
    # Находим максимальный существующий числовой код
    max_code = 0
    for code in existing_codes:
        try:
            num = int(code)
            if num > max_code:
                max_code = num
        except (ValueError, TypeError):
            pass
    
    # Начинаем с max_code + 1
    idx = max_code + 1
    while True:
        candidate = f"{idx:05d}"
        if candidate not in existing_codes:
            return candidate
        idx += 1


def _get_part_type_id_by_name(db: Session, name: str) -> int | None:
    """Find part_type_id by name (case-insensitive). Returns default (12) if not found."""
    if not name:
        return 12  # Default to "Тормозная система"
    from app.models.part_type import PartType
    pt = db.query(PartType).filter(
        PartType.name.ilike(name.strip())
    ).first()
    return pt.id if pt else 12  # Default to "Тормозная система" if not found


def _has_avito_integration(db: Session, org_id: str) -> bool:
    row = db.query(OrganizationAvitoIntegration).filter(
        OrganizationAvitoIntegration.organization_id == org_id
    ).first()
    if not row:
        return False
    if not row.enabled:
        return False
    return bool(row.client_id and row.client_secret_encrypted and row.avito_user_id)


def _job_to_response(job: AvitoAutoloadJob) -> AvitoAutoloadJobResponse:
    return AvitoAutoloadJobResponse(
        id=job.id,
        job_type=job.job_type,
        status=job.status,
        stage=job.stage,
        processed_count=job.processed_count,
        total_count=job.total_count,
        result_file_ref=job.result_file_ref,
        error_summary=job.error_summary,
        result=_json_loads(job.result_json, None),
        created_at=job.created_at.isoformat() if job.created_at else None,
        updated_at=job.updated_at.isoformat() if job.updated_at else None,
    )


def _resolve_saved_autoload_file(db: Session, org_id: str) -> tuple[Path, str]:
    base_dir = Path(__file__).resolve().parents[2] / "uploads" / "avito" / org_id
    base_dir.mkdir(parents=True, exist_ok=True)
    _remove_existing_avito_xlsx_files(base_dir)
    dest_name = "autoload.xlsx"
    xlsx_path = base_dir / dest_name
    rel_path = f"/uploads/avito/{org_id}/{dest_name}"
    return xlsx_path, rel_path


async def _push_file_to_avito(db: Session, org_id: str, filename: str, file_bytes: bytes) -> tuple[Any, Any, Any, Any]:
    row = db.query(OrganizationAvitoIntegration).filter(
        OrganizationAvitoIntegration.organization_id == org_id
    ).first()
    avito_upload = None
    avito_upload_status = None
    avito_report = None
    avito_token_error = None
    if not row:
        return avito_upload, avito_upload_status, avito_report, avito_token_error

    try:
        sec = decrypt_secret(row.client_secret_encrypted)
        token = await avito_api_svc.fetch_access_token(row.client_id, sec)
        avito_upload_status, avito_upload = await avito_api_svc.upload_autoload_xlsx(
            token, filename, file_bytes
        )
        if avito_upload_status in (200, 201):
            avito_report = await avito_api_svc.get_last_completed_report_v3(token)
            if avito_report is None:
                avito_report = await avito_api_svc.get_last_report_v1(
                    token, int(row.avito_user_id)
                )
    except Exception as e:
        logger.exception("Avito API error")
        avito_token_error = str(e)

    return avito_upload, avito_upload_status, avito_report, avito_token_error


@router.get("/{org_id}/avito/autoload/category-tree", response_model=AvitoAutoloadCategoryTreeResponse)
async def get_avito_autoload_category_tree(
    org_id: str,
    db: Session = Depends(get_db),
    current_user: UserModel = Depends(get_current_user),
):
    _ensure_org_access(current_user, org_id)
    _org_exists(db, org_id)

    now = time.time()
    cached = _CATEGORY_TREE_CACHE.get(org_id)
    if cached and now - cached[0] < _CATEGORY_TREE_TTL_S:
        raw = cached[1]
        return AvitoAutoloadCategoryTreeResponse(tree=_normalize_category_tree(raw), raw=raw)

    row = db.query(OrganizationAvitoIntegration).filter(
        OrganizationAvitoIntegration.organization_id == org_id
    ).first()
    if not row or not row.client_id or not row.client_secret_encrypted:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Интеграция Авито не настроена")

    try:
        sec = decrypt_secret(row.client_secret_encrypted)
        token = await avito_api_svc.fetch_access_token(row.client_id, sec)
        raw = await avito_api_svc.get_autoload_user_docs_tree(token)
    except Exception as e:
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail=str(e))

    _CATEGORY_TREE_CACHE[org_id] = (now, raw)
    return AvitoAutoloadCategoryTreeResponse(tree=_normalize_category_tree(raw), raw=raw)


@router.get("/{org_id}/avito/credentials", response_model=AvitoCredentialsResponse)
def get_avito_credentials(
    org_id: str,
    db: Session = Depends(get_db),
    current_user: UserModel = Depends(get_current_user),
):
    _ensure_org_access(current_user, org_id)
    _org_exists(db, org_id)
    row = db.query(OrganizationAvitoIntegration).filter(
        OrganizationAvitoIntegration.organization_id == org_id
    ).first()
    last = _get_last_autoload(db, org_id)
    if not row:
        return AvitoCredentialsResponse(
            client_id="",
            avito_user_id=None,
            client_secret_configured=False,
            enabled=True,
            last_autoload=last,
        )
    return AvitoCredentialsResponse(
        client_id=row.client_id,
        avito_user_id=int(row.avito_user_id),
        client_secret_configured=True,
        enabled=row.enabled,
        last_autoload=last,
    )


@router.post("/{org_id}/avito/autoload/export-async", response_model=AvitoAutoloadJobResponse)
async def export_products_to_avito_autoload_async(
    org_id: str,
    body: AvitoAutoloadExportAsyncRequest,
    db: Session = Depends(get_db),
    current_user: UserModel = Depends(get_current_user),
):
    _ensure_org_access(current_user, org_id)
    _org_exists(db, org_id)
    if not _has_avito_integration(db, org_id):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Интеграция Авито не настроена")

    requested_ids = list(dict.fromkeys(body.product_ids))
    if not requested_ids:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Список product_ids пуст")

    job = AvitoAutoloadJob(
        organization_id=org_id,
        created_by=current_user.id,
        job_type="export",
        status="pending",
        stage="queued",
        processed_count=0,
        total_count=len(requested_ids),
        payload_json=json.dumps(
            {
                "product_ids": requested_ids,
                "publish_after_export": bool(body.publish_after_export),
            },
            ensure_ascii=False,
        ),
    )
    db.add(job)
    db.commit()
    db.refresh(job)

    task = run_avito_export_job.delay(job.id)
    job.celery_task_id = task.id
    db.commit()
    db.refresh(job)
    return _job_to_response(job)


@router.post("/{org_id}/avito/autoload/publish-async", response_model=AvitoAutoloadJobResponse)
async def publish_avito_autoload_async(
    org_id: str,
    db: Session = Depends(get_db),
    current_user: UserModel = Depends(get_current_user),
):
    _ensure_org_access(current_user, org_id)
    _org_exists(db, org_id)
    cache = db.query(OrganizationAvitoAutoloadCache).filter(
        OrganizationAvitoAutoloadCache.organization_id == org_id
    ).first()
    if not cache or not cache.saved_path:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Файл автозагрузки не найден")
    job = AvitoAutoloadJob(
        organization_id=org_id,
        created_by=current_user.id,
        job_type="publish",
        status="pending",
        stage="queued",
        processed_count=0,
        total_count=1,
    )
    db.add(job)
    db.commit()
    db.refresh(job)

    task = run_avito_publish_job.delay(job.id)
    job.celery_task_id = task.id
    db.commit()
    db.refresh(job)
    return _job_to_response(job)


@router.get("/{org_id}/avito/autoload/jobs/{job_id}", response_model=AvitoAutoloadJobResponse)
async def get_avito_autoload_job_status(
    org_id: str,
    job_id: int,
    db: Session = Depends(get_db),
    current_user: UserModel = Depends(get_current_user),
):
    _ensure_org_access(current_user, org_id)
    job = db.query(AvitoAutoloadJob).filter(
        AvitoAutoloadJob.id == job_id,
        AvitoAutoloadJob.organization_id == org_id,
    ).first()
    if not job:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Job не найден")
    return _job_to_response(job)


@router.post("/{org_id}/avito/autoload/export", response_model=AvitoAutoloadExportResponse)
async def export_products_to_avito_autoload(
    org_id: str,
    body: AvitoAutoloadExportRequest,
    db: Session = Depends(get_db),
    current_user: UserModel = Depends(get_current_user),
):
    _ensure_org_access(current_user, org_id)
    org = _org_exists(db, org_id)
    if not _has_avito_integration(db, org_id):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Интеграция Авито не настроена")

    requested_ids = list(dict.fromkeys(body.product_ids))
    products = (
        db.query(ProductModel)
        .options(selectinload(ProductModel.part_type))
        .filter(
            ProductModel.organization_id == org_id,
            ProductModel.id.in_(requested_ids),
        )
        .all()
    )
    by_id = {p.id: p for p in products}
    storage_ids = {p.storage_location_id for p in products if p.storage_location_id}
    storage_rows = (
        db.query(StorageLocationModel).filter(StorageLocationModel.id.in_(list(storage_ids))).all()
        if storage_ids
        else []
    )
    storage_by_id = {s.id: s for s in storage_rows}
    export_rows: list[dict[str, Any]] = []
    for product_id in requested_ids:
        product = by_id.get(product_id)
        if not product:
            continue
        raw_photos = [ph.photo_url for ph in (product.photos or []) if ph.photo_url]
        photos = await ensure_local_pictures(
            raw_photos,
            org_id=org_id,
            db=db,
            for_xlsx=True,
            limit=5,
            soft_fail=True,
            per_photo_timeout_s=25.0,
            celery_timeout_s=120,
        )
        avito_link = db.query(ProductAvitoListingLink).filter(
            ProductAvitoListingLink.organization_id == org_id,
            ProductAvitoListingLink.product_id == product.id,
        ).first()
        # category больше не используется - всегда "Запчасти и аксессуары"
        storage = storage_by_id.get(product.storage_location_id) if product.storage_location_id else None
        address = (storage.address if storage and storage.address else None) or (org.address or "")
        export_rows.append(
            {
                "id": product.id,
                "internal_code": product.internal_code,
                "article": product.article,
                "brand": product.brand,
                "is_new": product.is_new,
                "price": product.price,
                "name": product.name,
                "description": product.description,
                "quantity": product.quantity,
                "photos": photos,
                "avito_id": avito_link.avito_ad_id if avito_link else "",
                "category": "Запчасти и аксессуары",  # Всегда эта категория для листа "Объявления"
                "template_sheet": "Объявления",  # Всегда этот лист
                "address": address,
                "part_type_name": product.part_type.name if product.part_type else "",
                # NEW: Map to new format fields
                "availability": "В наличии" if product.quantity > 0 else "Под заказ",
                "originality": "Оригинал" if product.is_new else "Неоригинал",
            }
        )
    if not export_rows:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Товары не найдены для экспорта")

    xlsx_path, rel_path = _resolve_saved_autoload_file(db, org_id)
    existing_bytes = xlsx_path.read_bytes() if xlsx_path.is_file() else None
    merged_bytes = upsert_products_to_avito_autoload(
        existing_bytes,
        export_rows,
        public_base_url=(settings.PUBLIC_BASE_URL or settings.BASE_URL or "").strip(),
    )
    xlsx_path.write_bytes(merged_bytes)
    parsed = parse_and_validate_avito_autoload(merged_bytes)
    _save_autoload_cache(
        db,
        org_id,
        saved_path=rel_path,
        items=parsed.items,
        local_validation_ok=parsed.local_ok,
        local_errors=parsed.local_errors,
        sheets_parsed=parsed.sheets_parsed,
        avito_upload=None,
        avito_upload_status=None,
        avito_report=None,
        avito_token_error=None,
    )
    return AvitoAutoloadExportResponse(
        saved_path=rel_path,
        items=parsed.items,
        local_validation_ok=parsed.local_ok,
        local_errors=parsed.local_errors,
        sheets_parsed=parsed.sheets_parsed,
        exported_count=len(export_rows),
    )


@router.post("/{org_id}/avito/autoload/export/{product_id}", response_model=AvitoAutoloadExportResponse)
async def export_single_product_to_avito_autoload(
    org_id: str,
    product_id: int,
    db: Session = Depends(get_db),
    current_user: UserModel = Depends(get_current_user),
):
    return await export_products_to_avito_autoload(
        org_id=org_id,
        body=AvitoAutoloadExportRequest(product_ids=[product_id]),
        db=db,
        current_user=current_user,
    )


@router.put("/{org_id}/avito/credentials", response_model=AvitoCredentialsResponse)
def put_avito_credentials(
    org_id: str,
    body: AvitoCredentialsUpdate,
    db: Session = Depends(get_db),
    current_user: UserModel = Depends(get_current_user),
):
    _ensure_org_access(current_user, org_id)
    _org_exists(db, org_id)

    row = db.query(OrganizationAvitoIntegration).filter(
        OrganizationAvitoIntegration.organization_id == org_id
    ).first()

    secret_in = (body.client_secret or "").strip()
    if row is None:
        if not secret_in:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="При первом сохранении укажите client_secret",
            )
        row = OrganizationAvitoIntegration(
            organization_id=org_id,
            avito_user_id=body.avito_user_id,
            client_id=body.client_id.strip(),
            client_secret_encrypted=encrypt_secret(secret_in),
            enabled=True,
        )
        db.add(row)
    else:
        row.client_id = body.client_id.strip()
        row.avito_user_id = body.avito_user_id
        if secret_in:
            row.client_secret_encrypted = encrypt_secret(secret_in)

    db.commit()
    db.refresh(row)
    last = _get_last_autoload(db, org_id)
    return AvitoCredentialsResponse(
        client_id=row.client_id,
        avito_user_id=int(row.avito_user_id),
        client_secret_configured=True,
        enabled=row.enabled,
        last_autoload=last,
    )


@router.patch("/{org_id}/avito/toggle-enabled", response_model=AvitoCredentialsResponse)
def toggle_avito_integration_enabled(
    org_id: str,
    db: Session = Depends(get_db),
    current_user: UserModel = Depends(get_current_user),
):
    """Toggle Avito integration enabled/disabled state."""
    _ensure_org_access(current_user, org_id)
    _org_exists(db, org_id)

    row = db.query(OrganizationAvitoIntegration).filter(
        OrganizationAvitoIntegration.organization_id == org_id
    ).first()

    if not row:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Интеграция Авито не настроена. Сначала подключите API.",
        )

    # Toggle the enabled state
    row.enabled = not row.enabled
    db.commit()
    db.refresh(row)

    last = _get_last_autoload(db, org_id)
    return AvitoCredentialsResponse(
        client_id=row.client_id,
        avito_user_id=int(row.avito_user_id),
        client_secret_configured=True,
        enabled=row.enabled,
        last_autoload=last,
    )


@router.post("/{org_id}/avito/autoload/upload", response_model=AvitoAutoloadUploadResponse)
async def upload_avito_autoload(
    org_id: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: UserModel = Depends(get_current_user),
):
    _ensure_org_access(current_user, org_id)
    _org_exists(db, org_id)

    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Ожидается файл .xlsx")

    body = await file.read()
    if len(body) > 50 * 1024 * 1024:
        raise HTTPException(status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, detail="Файл больше 50MB")

    base_dir = Path(__file__).resolve().parents[2] / "uploads" / "avito" / org_id
    base_dir.mkdir(parents=True, exist_ok=True)
    _remove_existing_avito_xlsx_files(base_dir)

    dest_name = "autoload.xlsx"
    dest_path = base_dir / dest_name
    dest_path.write_bytes(body)

    # Добавляем лист "Объявления" из template.xlsx если его ещё нет
    try:
        template_path = Path(__file__).resolve().parents[2] / "templates" / "avito" / "template.xlsx"
        if template_path.is_file():
            # Открываем загруженный файл
            wb = load_workbook(str(dest_path), read_only=False)
            
            # Открываем template
            wb_template = load_workbook(str(template_path), read_only=False)
            
            # Проверяем, есть ли уже лист "Объявления"
            if "Объявления" not in wb.sheetnames:
                # Копируем лист из template
                if "Объявления" in wb_template.sheetnames:
                    source_sheet = wb_template["Объявления"]
                    
                    # Создаём новый лист в файле пользователя
                    target_sheet = wb.create_sheet("Объявления")
                    
                    # Копируем все данные из template
                    for row in source_sheet.iter_rows():
                        for cell in row:
                            target_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                            
                    # Копируем настройки колонок (ширину и т.д.)
                    if hasattr(source_sheet, 'column_dimensions'):
                        for col_dim in source_sheet.column_dimensions:
                            target_sheet.column_dimensions[col_dim] = source_sheet.column_dimensions[col_dim]
                    
                    wb.save(str(dest_path))
                    print(f"✅ Added 'Объявления' sheet from template to {dest_path}")
            
            wb.close()
            wb_template.close()
    except Exception as e:
        print(f"⚠️ Warning: Could not add 'Объявления' sheet: {e}")
        # Продолжаем даже если не удалось добавить лист

    # Перечитываем файл после добавления листа "Объявления"
    updated_body = dest_path.read_bytes()
    parsed = parse_and_validate_avito_autoload(updated_body)

    # Debug logging for XLSX parsing
    print(f"📊 Avito XLSX Parse Results:")
    print(f"  - Items found: {len(parsed.items)}")
    print(f"  - Sheets parsed: {parsed.sheets_parsed}")
    print(f"  - Local OK: {parsed.local_ok}")
    print(f"  - Local errors: {len(parsed.local_errors)}")
    if parsed.local_errors:
        for err in parsed.local_errors[:5]:  # First 5 errors
            print(f"    Error: {err}")

    # Проверка формата файла и предупреждения
    warnings = []
    if len(parsed.items) == 0 and not parsed.local_errors:
        warnings.append(
            "Файл загружен, но товары не найдены. "
            "Убедитесь, что файл содержит лист с колонкой 'Title' и данными товаров, "
            "начиная со второй строки."
        )
    
    if not parsed.sheets_parsed or len(parsed.sheets_parsed) == 0:
        warnings.append(
            "В файле не найдено листов с данными. "
            "Проверьте что файл содержит корректный шаблон Авито с английскими заголовками "
            "(Title, Price, Brand, OEM и т.д.)."
        )
    
    if parsed.local_ok and len(parsed.items) == 0:
        warnings.append(
            "Локальная проверка пройдена, но товаров нет. "
            "Возможно файл содержит только заголовки без данных."
        )

    # Создаем записи в product_avito_listing_links для товаров из файла
    # у которых есть AvitoId (уже опубликованы на Avito)
    try:
        avito_id_count = 0
        for item in parsed.items:
            # Проверяем есть ли у товара AvitoId
            avito_id = item.get('avito_id')
            
            # Для сопоставления с продуктом используем internal_code (уникальный ID)
            # который хранится в колонке "Id" в XLSX
            internal_code = item.get('unique_ad_id')
            
            if avito_id and internal_code:
                # Находим продукт по internal_code
                product = db.query(ProductModel).filter(
                    ProductModel.organization_id == org_id,
                    ProductModel.internal_code == str(internal_code),
                ).first()
                
                if product:
                    # Проверяем существует ли уже запись
                    existing_link = db.query(ProductAvitoListingLink).filter(
                        ProductAvitoListingLink.organization_id == org_id,
                        ProductAvitoListingLink.product_id == product.id,
                    ).first()
                    
                    if not existing_link:
                        # Создаем новую запись
                        new_link = ProductAvitoListingLink(
                            organization_id=org_id,
                            product_id=product.id,
                            avito_ad_id=str(avito_id),
                        )
                        db.add(new_link)
                        avito_id_count += 1
                        print(f"  ✅ Created Avito link for product {product.id} (internal_code: {internal_code})")
                    elif existing_link.avito_ad_id != str(avito_id):
                        # Обновляем Avito ID если он изменился
                        existing_link.avito_ad_id = str(avito_id)
                        avito_id_count += 1
                        print(f"  ✅ Updated Avito link for product {product.id} (internal_code: {internal_code})")
        
        if avito_id_count > 0:
            db.commit()
            print(f"✅ Created/updated {avito_id_count} Avito listing link(s) from nomenclature import")
    except Exception as e:
        logger.exception("Failed to create Avito listing links from nomenclature import")
        db.rollback()
        # Не прерываем загрузку файла из-за ошибок в links

    # Автопубликация отключена: загрузка файла только сохраняет и валидирует XLSX.
    avito_upload, avito_upload_status, avito_report, avito_token_error = (None, None, None, None)

    rel_path = f"/uploads/avito/{org_id}/{dest_name}"
    try:
        _save_autoload_cache(
            db,
            org_id,
            saved_path=rel_path,
            items=parsed.items,
            local_validation_ok=parsed.local_ok,
            local_errors=parsed.local_errors,
            sheets_parsed=parsed.sheets_parsed,
            avito_upload=avito_upload,
            avito_upload_status=avito_upload_status,
            avito_report=avito_report,
            avito_token_error=avito_token_error,
            warnings=warnings if warnings else None,
        )
    except Exception:
        logger.exception("Не удалось сохранить кэш автозагрузки")
        db.rollback()

    return AvitoAutoloadUploadResponse(
        saved_path=rel_path,
        items=parsed.items,
        local_validation_ok=parsed.local_ok,
        local_errors=parsed.local_errors,
        sheets_parsed=parsed.sheets_parsed,
        avito_upload=avito_upload,
        avito_upload_status=avito_upload_status,
        avito_report=avito_report,
        avito_token_error=avito_token_error,
        warnings=warnings if warnings else None,
    )


@router.post("/{org_id}/avito/autoload/set-category", response_model=AvitoAutoloadUploadResponse)
async def set_avito_autoload_category(
    org_id: str,
    body: AvitoAutoloadSetCategoryRequest,
    db: Session = Depends(get_db),
    current_user: UserModel = Depends(get_current_user),
):
    _ensure_org_access(current_user, org_id)
    _org_exists(db, org_id)

    cache = db.query(OrganizationAvitoAutoloadCache).filter(
        OrganizationAvitoAutoloadCache.organization_id == org_id
    ).first()
    if not cache or not cache.saved_path:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Файл автозагрузки не найден")

    rel_path = cache.saved_path
    project_root = Path(__file__).resolve().parents[2]
    xlsx_path = project_root / rel_path.lstrip("/")
    if not xlsx_path.is_file():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Файл автозагрузки на диске не найден")

    wb = load_workbook(str(xlsx_path), read_only=False)
    try:
        if body.sheet not in wb.sheetnames:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Лист '{body.sheet}' не найден")
        ws = wb[body.sheet]
        category_col = None
        # Не используем ws[2] — в некоторых файлах openpyxl может отдавать усечённый ряд.
        # Сканируем явными cell() до max_column.
        for col_idx in range(1, (ws.max_column or 0) + 1):
            h = str(ws.cell(row=2, column=col_idx).value or "").strip()
            if h == "Категория":
                category_col = col_idx
                break
        if not category_col:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"На листе '{body.sheet}' не найдена колонка 'Категория'",
            )
        ws.cell(row=int(body.row), column=category_col, value="Запчасти и аксессуары")

        out = BytesIO()
        wb.save(out)
        updated_bytes = out.getvalue()
    finally:
        wb.close()

    xlsx_path.write_bytes(updated_bytes)
    parsed = parse_and_validate_avito_autoload(updated_bytes)
    avito_upload, avito_upload_status, avito_report, avito_token_error = (None, None, None, None)
    _save_autoload_cache(
        db,
        org_id,
        saved_path=rel_path,
        items=parsed.items,
        local_validation_ok=parsed.local_ok,
        local_errors=parsed.local_errors,
        sheets_parsed=parsed.sheets_parsed,
        avito_upload=avito_upload,
        avito_upload_status=avito_upload_status,
        avito_report=avito_report,
        avito_token_error=avito_token_error,
    )

    return AvitoAutoloadUploadResponse(
        saved_path=rel_path,
        items=parsed.items,
        local_validation_ok=parsed.local_ok,
        local_errors=parsed.local_errors,
        sheets_parsed=parsed.sheets_parsed,
        avito_upload=avito_upload,
        avito_upload_status=avito_upload_status,
        avito_report=avito_report,
        avito_token_error=avito_token_error,
    )


@router.post("/{org_id}/avito/autoload/set-ad-type", response_model=AvitoAutoloadUploadResponse)
async def set_avito_autoload_ad_type(
    org_id: str,
    body: AvitoAutoloadSetAdTypeRequest,
    db: Session = Depends(get_db),
    current_user: UserModel = Depends(get_current_user),
):
    _ensure_org_access(current_user, org_id)
    _org_exists(db, org_id)

    cache = db.query(OrganizationAvitoAutoloadCache).filter(
        OrganizationAvitoAutoloadCache.organization_id == org_id
    ).first()
    if not cache or not cache.saved_path:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Файл автозагрузки не найден")

    rel_path = cache.saved_path
    project_root = Path(__file__).resolve().parents[2]
    xlsx_path = project_root / rel_path.lstrip("/")
    if not xlsx_path.is_file():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Файл автозагрузки на диске не найден")

    normalized_ad_type = str(body.ad_type or "").strip()
    allowed_ad_types = {"", "Товар приобретен на продажу", "Товар от производителя"}
    if normalized_ad_type not in allowed_ad_types:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Некорректный вид объявления",
        )

    wb = load_workbook(str(xlsx_path), read_only=False)
    try:
        if body.sheet not in wb.sheetnames:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Лист '{body.sheet}' не найден")
        ws = wb[body.sheet]
        ad_type_col = None
        for col_idx in range(1, (ws.max_column or 0) + 1):
            h = str(ws.cell(row=2, column=col_idx).value or "").strip()
            if h == "Вид объявления":
                ad_type_col = col_idx
                break
        if not ad_type_col:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"На листе '{body.sheet}' не найдена колонка 'Вид объявления'",
            )

        ws.cell(row=int(body.row), column=ad_type_col, value="Товар приобретен на продажу")

        out = BytesIO()
        wb.save(out)
        updated_bytes = out.getvalue()
    finally:
        wb.close()

    xlsx_path.write_bytes(updated_bytes)
    parsed = parse_and_validate_avito_autoload(updated_bytes)
    avito_upload, avito_upload_status, avito_report, avito_token_error = (None, None, None, None)
    _save_autoload_cache(
        db,
        org_id,
        saved_path=rel_path,
        items=parsed.items,
        local_validation_ok=parsed.local_ok,
        local_errors=parsed.local_errors,
        sheets_parsed=parsed.sheets_parsed,
        avito_upload=avito_upload,
        avito_upload_status=avito_upload_status,
        avito_report=avito_report,
        avito_token_error=avito_token_error,
    )

    return AvitoAutoloadUploadResponse(
        saved_path=rel_path,
        items=parsed.items,
        local_validation_ok=parsed.local_ok,
        local_errors=parsed.local_errors,
        sheets_parsed=parsed.sheets_parsed,
        avito_upload=avito_upload,
        avito_upload_status=avito_upload_status,
        avito_report=avito_report,
        avito_token_error=avito_token_error,
    )


@router.post("/{org_id}/avito/autoload/import", response_model=AvitoAutoloadImportResponse)
async def import_avito_autoload_rows(
    org_id: str,
    body: AvitoAutoloadImportRequest,
    db: Session = Depends(get_db),
    current_user: UserModel = Depends(get_current_user),
):
    _ensure_org_access(current_user, org_id)
    _org_exists(db, org_id)

    if not current_user.is_seller:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Только продавцы могут импортировать товары")
    if not body.use_file_price and body.sale_price is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Укажите цену прихода или включите цену из файла")

    storage_location = db.query(StorageLocationModel).filter(
        StorageLocationModel.id == body.storage_location_id,
        StorageLocationModel.organization_id == org_id,
    ).first()
    if not storage_location:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Склад не найден или не принадлежит организации")

    cache = (
        db.query(OrganizationAvitoAutoloadCache)
        .filter(OrganizationAvitoAutoloadCache.organization_id == org_id)
        .first()
    )
    if not cache:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Нет данных автозагрузки для импорта")

    items = _json_loads(cache.items_json, [])
    index = {(str(i.get("sheet")), int(i.get("row"))): i for i in items if i.get("sheet") and i.get("row") is not None}
    row_keys = [(r.sheet, int(r.row)) for r in body.rows]

    created_products = 0
    updated_products = 0
    created_stock_ins = 0
    skipped_rows: list[dict[str, Any]] = []

    # Prepare access to saved XLSX for in-place rewrite of photo links (optional).
    xlsx_path = None
    rel_path = None
    wb = None
    photo_col_cache: dict[str, int] = {}

    for key in row_keys:
        item = index.get((key[0], key[1]))
        if not item:
            skipped_rows.append({"sheet": key[0], "row": key[1], "reason": "Строка не найдена в кэше автозагрузки"})
            continue

        avito_id = str(item.get("avito_id") or "").strip()
        unique_ad_id = str(item.get("unique_ad_id") or "").strip()

        link = None
        if avito_id:
            link = db.query(ProductAvitoListingLink).filter(
                ProductAvitoListingLink.organization_id == org_id,
                ProductAvitoListingLink.avito_ad_id == avito_id,
            ).first()

        file_price = None
        try:
            file_price = float(item.get("price")) if item.get("price") not in (None, "") else None
        except (TypeError, ValueError):
            file_price = None

        effective_price = file_price if body.use_file_price else body.sale_price
        if effective_price is None:
            effective_price = body.sale_price if body.sale_price is not None else 1.0

        title = str(item.get("title") or "").strip()
        part_number = str(item.get("part_number") or "").strip()
        manufacturer = str(item.get("manufacturer") or "").strip()
        description = str(item.get("description") or "").strip()
        file_quantity = None
        try:
            raw_qty = item.get("quantity")
            if raw_qty not in (None, ""):
                file_quantity = int(float(str(raw_qty).replace(",", ".").strip()))
        except (TypeError, ValueError):
            file_quantity = None
        effective_quantity = file_quantity if file_quantity and file_quantity > 0 else int(body.quantity)

        product = None
        if link:
            product = db.query(ProductModel).filter(
                ProductModel.id == link.product_id,
                ProductModel.organization_id == org_id,
            ).first()
            if product is None:
                link = None
                stale_link = db.query(ProductAvitoListingLink).filter(
                    ProductAvitoListingLink.organization_id == org_id,
                    ProductAvitoListingLink.avito_ad_id == avito_id,
                ).first()
                if stale_link:
                    db.delete(stale_link)
        if product is None and unique_ad_id:
            product = db.query(ProductModel).filter(
                ProductModel.organization_id == org_id,
                ProductModel.internal_code == unique_ad_id,
            ).first()

        if product is not None:
            # Update existing product
            product.article = (part_number or avito_id or unique_ad_id or f"ROW-{key[1]}")[:30]
            product.name = (title or part_number or f"Avito row {key[1]}")[:255]
            product.brand = (manufacturer or "Unknown")[:100]
            product.price = effective_price
            product.quantity = effective_quantity
            product.storage_location_id = body.storage_location_id
            # part_type_id is required - try to get from item or use default
            if item.get("part_type_id"):
                product.part_type_id = int(item.get("part_type_id"))
            elif item.get("part_type_name"):
                # Map part_type_name to part_type_id
                product.part_type_id = _get_part_type_id_by_name(db, item.get("part_type_name"))
            elif not product.part_type_id:
                product.part_type_id = 12  # Default to "Тормозная система"
            if description:
                product.description = description
            updated_products += 1

        if product is None:
            # Create new product
            internal_code = unique_ad_id or _next_internal_code(db)
            # part_type_id is REQUIRED - try to get from item or use default
            part_type_id = None
            if item.get("part_type_id"):
                try:
                    part_type_id = int(item.get("part_type_id"))
                except (TypeError, ValueError):
                    part_type_id = None
            if not part_type_id and item.get("part_type_name"):
                # Map part_type_name to part_type_id
                part_type_id = _get_part_type_id_by_name(db, item.get("part_type_name"))
            if not part_type_id:
                part_type_id = 12  # Default to "Тормозная система" (id=12)
            
            try:
                product = ProductModel(
                    article=(part_number or avito_id or internal_code or f"ROW-{key[1]}")[:30],
                    name=(title or part_number or f"Avito row {key[1]}")[:255],
                    brand=(manufacturer or "Unknown")[:100],
                    price=effective_price,
                    quantity=effective_quantity,
                    is_new=False,
                    internal_code=internal_code,
                    description=description or f"Imported from Avito autoload (ad_id={avito_id or 'n/a'})",
                    organization_id=org_id,
                    storage_location_id=body.storage_location_id,
                    created_by=current_user.id,
                    part_type_id=part_type_id,  # REQUIRED FIELD
                )
                db.add(product)
                db.flush()
                created_products += 1
            except IntegrityError:
                db.rollback()
                # Если internal_code уже существует, пробуем другой
                if not unique_ad_id:
                    internal_code = _next_internal_code(db)
                    product = ProductModel(
                        article=(part_number or avito_id or internal_code or f"ROW-{key[1]}")[:30],
                        name=(title or part_number or f"Avito row {key[1]}")[:255],
                        brand=(manufacturer or "Unknown")[:100],
                        price=effective_price,
                        quantity=effective_quantity,
                        is_new=False,
                        internal_code=internal_code,
                        description=description or f"Imported from Avito autoload (ad_id={avito_id or 'n/a'})",
                        organization_id=org_id,
                        storage_location_id=body.storage_location_id,
                        created_by=current_user.id,
                        part_type_id=part_type_id,  # REQUIRED FIELD
                    )
                    db.add(product)
                    db.flush()
                    created_products += 1
                else:
                    skipped_rows.append(
                        {
                            "sheet": key[0],
                            "row": key[1],
                            "reason": f"internal_code={internal_code} уже существует",
                        }
                    )
                    continue

        # Ensure link exists if AvitoId is provided
        if avito_id and (link is None):
            try:
                with db.begin_nested():
                    db.add(
                        ProductAvitoListingLink(
                            organization_id=org_id,
                            product_id=product.id,
                            avito_ad_id=avito_id,
                        )
                    )
                    db.flush()
            except IntegrityError:
                # Another product may already have this AvitoId; don't fail whole import.
                skipped_rows.append(
                    {
                        "sheet": key[0],
                        "row": key[1],
                        "reason": f"Связь уже существует для AvitoId={avito_id}",
                    }
                )

        db.query(ProductPhoto).filter(ProductPhoto.product_id == product.id).delete()
        db.query(ProductVideo).filter(ProductVideo.product_id == product.id).delete()

        photos = item.get("photos") or []
        videos = item.get("videos") or []
        if isinstance(photos, list) and photos:
            processed_paths = await ensure_local_pictures(
                photos,
                org_id=org_id,
                db=db,
                for_xlsx=False,
                limit=5,
                soft_fail=True,
                per_photo_timeout_s=25.0,
                celery_timeout_s=120,
            )
            processed_urls_for_xlsx = [normalize_for_xlsx(p) for p in processed_paths]

            for p_url in processed_paths[:5]:
                p_url = str(p_url).strip()
                if not p_url:
                    continue
                db.add(
                    ProductPhoto(
                        product_id=product.id,
                        photo_url=p_url,
                        organization_id=org_id,
                        processing_status="completed",
                    )
                )

            # Rewrite XLSX cell 'Ссылки на фото' for this row so UI/next visits see local URLs.
            try:
                cache_row = db.query(OrganizationAvitoAutoloadCache).filter(
                    OrganizationAvitoAutoloadCache.organization_id == org_id
                ).first()
                if cache_row and cache_row.saved_path:
                    rel_path = cache_row.saved_path
                    xlsx_path = Path(__file__).resolve().parents[2] / rel_path.lstrip("/")
                    if xlsx_path.is_file():
                        if wb is None:
                            wb = load_workbook(str(xlsx_path), read_only=False)
                        sheet_name = str(key[0])
                        row_no = int(key[1])
                        if sheet_name in wb.sheetnames:
                            ws = wb[sheet_name]
                            if sheet_name not in photo_col_cache:
                                col = None
                                for col_idx in range(1, (ws.max_column or 0) + 1):
                                    # НОВЫЙ ФОРМАТ: заголовки в Row 1
                                    h = str(ws.cell(row=1, column=col_idx).value or "").strip()
                                    if h in ("ImageUrls", "Ссылки на фото", "Фото"):
                                        col = col_idx
                                        break
                                if col:
                                    photo_col_cache[sheet_name] = col
                            col = photo_col_cache.get(sheet_name)
                            if col:
                                ws.cell(row=row_no, column=col, value=" | ".join(processed_urls_for_xlsx))
            except Exception:
                # Soft fail: DB import should still succeed.
                pass
        if isinstance(videos, list):
            for v in videos[:1]:
                v_url = str(v).strip()
                if not v_url:
                    continue
                db.add(
                    ProductVideo(
                        product_id=product.id,
                        video_url=v_url,
                        organization_id=org_id,
                        processing_status="completed",
                    )
                )

        db.add(
            StockInModel(
                quantity=effective_quantity,
                sale_price=effective_price,
                organization_id=org_id,
                storage_location_id=body.storage_location_id,
                product_id=product.id,
                created_by=current_user.id,
            )
        )
        created_stock_ins += 1

    # If we updated XLSX in-memory, persist it and refresh cache.
    if wb is not None and xlsx_path is not None and rel_path is not None:
        try:
            out = BytesIO()
            wb.save(out)
            wb.close()
            updated_bytes = out.getvalue()
            xlsx_path.write_bytes(updated_bytes)
            parsed = parse_and_validate_avito_autoload(updated_bytes)
            _save_autoload_cache(
                db,
                org_id,
                saved_path=rel_path,
                items=parsed.items,
                local_validation_ok=parsed.local_ok,
                local_errors=parsed.local_errors,
                sheets_parsed=parsed.sheets_parsed,
                avito_upload=None,
                avito_upload_status=None,
                avito_report=None,
                avito_token_error=None,
            )
        except Exception:
            try:
                wb.close()
            except Exception:
                pass

    db.commit()
    return AvitoAutoloadImportResponse(
        created_products=created_products,
        updated_products=updated_products,
        created_stock_ins=created_stock_ins,
        skipped_rows=skipped_rows,
    )


@router.post("/{org_id}/avito/autoload/publish", response_model=AvitoAutoloadPublishResponse)
async def publish_avito_autoload(
    org_id: str,
    db: Session = Depends(get_db),
    current_user: UserModel = Depends(get_current_user),
):
    _ensure_org_access(current_user, org_id)
    _org_exists(db, org_id)
    if not _has_avito_integration(db, org_id):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Интеграция Авито не настроена")

    cache = (
        db.query(OrganizationAvitoAutoloadCache)
        .filter(OrganizationAvitoAutoloadCache.organization_id == org_id)
        .first()
    )
    if not cache or not cache.saved_path:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Файл автозагрузки не найден")

    rel_path = cache.saved_path
    xlsx_path = Path(__file__).resolve().parents[2] / rel_path.lstrip("/")
    if not xlsx_path.is_file():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Файл автозагрузки на диске не найден")

    file_bytes = xlsx_path.read_bytes()
    avito_upload, avito_upload_status, avito_report, avito_token_error = await _push_file_to_avito(
        db, org_id, xlsx_path.name, file_bytes
    )
    _save_autoload_cache(
        db,
        org_id,
        saved_path=rel_path,
        items=_json_loads(cache.items_json, []),
        local_validation_ok=bool(cache.local_validation_ok),
        local_errors=_json_loads(cache.local_errors_json, []),
        sheets_parsed=_json_loads(cache.sheets_parsed_json, []),
        avito_upload=avito_upload,
        avito_upload_status=avito_upload_status,
        avito_report=avito_report,
        avito_token_error=avito_token_error,
    )
    return AvitoAutoloadPublishResponse(
        saved_path=rel_path,
        avito_upload=avito_upload,
        avito_upload_status=avito_upload_status,
        avito_report=avito_report,
        avito_token_error=avito_token_error,
    )


@router.post("/{org_id}/avito/autoload/actions", response_model=AvitoAutoloadUploadResponse)
async def apply_avito_autoload_actions(
    org_id: str,
    body: AvitoAutoloadApplyActionRequest,
    db: Session = Depends(get_db),
    current_user: UserModel = Depends(get_current_user),
):
    _ensure_org_access(current_user, org_id)
    _org_exists(db, org_id)

    cache = (
        db.query(OrganizationAvitoAutoloadCache)
        .filter(OrganizationAvitoAutoloadCache.organization_id == org_id)
        .first()
    )
    if not cache or not cache.saved_path:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Файл автозагрузки не найден")

    rel_path = cache.saved_path
    project_root = Path(__file__).resolve().parents[2]
    xlsx_path = project_root / rel_path.lstrip("/")
    if not xlsx_path.is_file():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Файл автозагрузки на диске не найден")

    target_value = ACTION_VALUES.get(body.action)
    if not target_value:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Неизвестное действие")

    wb = load_workbook(str(xlsx_path), read_only=False)
    row_keys = {(r.sheet, int(r.row)) for r in body.rows}
    updated_rows = 0
    missing_sheets: set[str] = set()

    for sheet_name, row_no in row_keys:
        if sheet_name not in wb.sheetnames:
            missing_sheets.add(sheet_name)
            continue
        ws = wb[sheet_name]
        header = [str(v).strip() if v is not None else "" for v in ws[2]]
        action_col = None
        for idx, h in enumerate(header, start=1):
            if h in ACTION_HEADERS:
                action_col = idx
                break
        if not action_col:
            wb.close()
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"На листе '{sheet_name}' не найдена колонка 'Действие'",
            )
        ws.cell(row=row_no, column=action_col, value=target_value)
        updated_rows += 1

    if missing_sheets:
        wb.close()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Не найдены листы: {', '.join(sorted(missing_sheets))}",
        )
    if updated_rows == 0:
        wb.close()
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Нет строк для обновления")

    out = BytesIO()
    wb.save(out)
    wb.close()
    updated_bytes = out.getvalue()
    xlsx_path.write_bytes(updated_bytes)

    parsed = parse_and_validate_avito_autoload(updated_bytes)
    # Автопубликация отключена: действия только обновляют файл.
    avito_upload, avito_upload_status, avito_report, avito_token_error = (None, None, None, None)

    _save_autoload_cache(
        db,
        org_id,
        saved_path=rel_path,
        items=parsed.items,
        local_validation_ok=parsed.local_ok,
        local_errors=parsed.local_errors,
        sheets_parsed=parsed.sheets_parsed,
        avito_upload=avito_upload,
        avito_upload_status=avito_upload_status,
        avito_report=avito_report,
        avito_token_error=avito_token_error,
    )

    return AvitoAutoloadUploadResponse(
        saved_path=rel_path,
        items=parsed.items,
        local_validation_ok=parsed.local_ok,
        local_errors=parsed.local_errors,
        sheets_parsed=parsed.sheets_parsed,
        avito_upload=avito_upload,
        avito_upload_status=avito_upload_status,
        avito_report=avito_report,
        avito_token_error=avito_token_error,
    )


@router.post("/{org_id}/avito/autoload/remove-rows", response_model=AvitoAutoloadUploadResponse)
async def remove_avito_autoload_rows(
    org_id: str,
    body: AvitoAutoloadRemoveRowsRequest,
    db: Session = Depends(get_db),
    current_user: UserModel = Depends(get_current_user),
):
    """Удаляет строки из XLSX файла и обновляет кэш."""
    _ensure_org_access(current_user, org_id)
    _org_exists(db, org_id)

    cache = (
        db.query(OrganizationAvitoAutoloadCache)
        .filter(OrganizationAvitoAutoloadCache.organization_id == org_id)
        .first()
    )
    if not cache or not cache.saved_path:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Файл автозагрузки не найден")

    rel_path = cache.saved_path
    project_root = Path(__file__).resolve().parents[2]
    xlsx_path = project_root / rel_path.lstrip("/")
    if not xlsx_path.is_file():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Файл автозагрузки на диске не найден")

    wb = load_workbook(str(xlsx_path), read_only=False)
    
    # Группируем строки по листам
    rows_by_sheet: dict[str, list[int]] = {}
    for row_info in body.rows:
        sheet_name = row_info.sheet
        row_no = int(row_info.row)
        if sheet_name not in rows_by_sheet:
            rows_by_sheet[sheet_name] = []
        rows_by_sheet[sheet_name].append(row_no)
    
    removed_count = 0
    missing_sheets: set[str] = set()
    
    for sheet_name, row_numbers in rows_by_sheet.items():
        if sheet_name not in wb.sheetnames:
            missing_sheets.add(sheet_name)
            continue
        
        ws = wb[sheet_name]
        # Сортируем номера строк по убыванию, чтобы удалять с конца (чтобы номера строк не смещались)
        row_numbers_sorted = sorted(row_numbers, reverse=True)
        
        for row_no in row_numbers_sorted:
            ws.delete_rows(row_no, 1)
            removed_count += 1
    
    if missing_sheets:
        wb.close()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Не найдены листы: {', '.join(sorted(missing_sheets))}",
        )
    if removed_count == 0:
        wb.close()
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Нет строк для удаления")

    out = BytesIO()
    wb.save(out)
    wb.close()
    updated_bytes = out.getvalue()
    xlsx_path.write_bytes(updated_bytes)

    parsed = parse_and_validate_avito_autoload(updated_bytes)
    avito_upload, avito_upload_status, avito_report, avito_token_error = (None, None, None, None)

    _save_autoload_cache(
        db,
        org_id,
        saved_path=rel_path,
        items=parsed.items,
        local_validation_ok=parsed.local_ok,
        local_errors=parsed.local_errors,
        sheets_parsed=parsed.sheets_parsed,
        avito_upload=avito_upload,
        avito_upload_status=avito_upload_status,
        avito_report=avito_report,
        avito_token_error=avito_token_error,
    )

    return AvitoAutoloadUploadResponse(
        saved_path=rel_path,
        items=parsed.items,
        local_validation_ok=parsed.local_ok,
        local_errors=parsed.local_errors,
        sheets_parsed=parsed.sheets_parsed,
        avito_upload=avito_upload,
        avito_upload_status=avito_upload_status,
        avito_report=avito_report,
        avito_token_error=avito_token_error,
    )
