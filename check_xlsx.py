#!/usr/bin/env python3
"""Check XLSX file and count items"""
from openpyxl import load_workbook

file_path = "c:/Users/khram/OneDrive/Рабочий стол/autoparts/test_file.xlsx"

print(f"📂 Opening file: {file_path}\n")

wb = load_workbook(file_path, read_only=False, data_only=True)

print(f"📊 Sheets found: {wb.sheetnames}\n")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"{'='*80}")
    print(f"📋 Sheet: '{sheet_name}'")
    print(f"   Max rows: {ws.max_row}")
    print(f"   Max columns: {ws.max_column}")
    
    # Get headers (Row 1)
    headers = []
    for i in range(1, min(30, ws.max_column + 1)):
        val = ws.cell(row=1, column=i).value
        if val:
            headers.append((i, str(val).strip()))
    
    print(f"\n📝 Headers (first 30 cols):")
    for col_num, header in headers:
        print(f"   Col {col_num}: {header}")
    
    # Count non-empty rows
    data_rows = 0
    empty_rows = 0
    
    # Check first few rows
    print(f"\n🔍 First 5 rows data:")
    for row_idx in range(2, min(7, ws.max_row + 1)):
        row_data = []
        for col_num, header in headers[:10]:  # First 10 headers
            val = ws.cell(row=row_idx, column=col_num).value
            row_data.append(str(val) if val else "")
        
        has_data = any(row_data)
        if has_data:
            data_rows += 1
            print(f"   Row {row_idx}: {' | '.join(row_data[:5])}...")
        else:
            empty_rows += 1
            print(f"   Row {row_idx}: [EMPTY]")
    
    # Count all rows with data
    total_data_rows = 0
    for row_idx in range(2, ws.max_row + 1):
        has_data = False
        for col_num, _ in headers:
            val = ws.cell(row=row_idx, column=col_num).value
            if val and str(val).strip():
                has_data = True
                break
        if has_data:
            total_data_rows += 1
    
    print(f"\n✅ Total data rows (excluding header): {total_data_rows}")
    print()

wb.close()

# Now test with actual parser
print(f"\n{'='*80}")
print(f"🧪 Testing with Avito parser...\n")

try:
    import sys
    sys.path.insert(0, 'c:/Users/khram/OneDrive/Рабочий стол/autoparts/backend')
    from app.services.avito_autoload_xlsx import parse_and_validate_avito_autoload
    
    with open(file_path, 'rb') as f:
        file_bytes = f.read()
    
    result = parse_and_validate_avito_autoload(file_bytes)
    
    print(f"\n📊 Parse Results:")
    print(f"   Items found: {len(result.items)}")
    print(f"   Sheets parsed: {result.sheets_parsed}")
    print(f"   Local OK: {result.local_ok}")
    print(f"   Local errors: {len(result.local_errors)}")
    
    if result.local_errors:
        print(f"\n⚠️  First 5 errors:")
        for err in result.local_errors[:5]:
            print(f"   - {err}")
    
    if result.items:
        print(f"\n📦 First 3 items:")
        for item in result.items[:3]:
            print(f"   - Row {item.get('row')}: {item.get('title')} | Price: {item.get('price')} | Brand: {item.get('manufacturer')}")
        
        if len(result.items) > 3:
            print(f"   ... and {len(result.items) - 3} more items")

except Exception as e:
    print(f"❌ Error parsing: {e}")
    import traceback
    traceback.print_exc()
